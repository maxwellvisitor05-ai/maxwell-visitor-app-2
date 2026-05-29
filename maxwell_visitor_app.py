#!/usr/bin/env python3
"""Maxwell Engineering Solutions - Visitor Management System v3.0"""

from flask import Flask, request, jsonify, redirect, session, send_file
import sqlite3, base64, io, os, hashlib, json
from datetime import datetime, timezone, timedelta

try:
    import qrcode
    HAS_QR = True
except:
    HAS_QR = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_EXCEL = True
except:
    HAS_EXCEL = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'maxwell2024secret')
app.permanent_session_lifetime = __import__('datetime').timedelta(days=30)

@app.before_request
def make_session_permanent():
    session.permanent = True

DB = "maxwell_visitors.db"
PANTRY_EMAIL = os.environ.get("PANTRY_EMAIL","maxwellvisitor05@gmail.com")
PANTRY_PASSWORD = os.environ.get("PANTRY_PASSWORD","MaxwellPantry2024")
SECURITY_MOBILE = "9023730509"

EMPLOYEE_EMAILS = {
    "Nishit Patel":"production@maxwells.in","Vaibhav Desai":"qa@maxwells.in",
    "Vasant Sarla":"qc@maxwells.in","Mohit Goswami":"head.hoa@maxwells.in",
    "Vrunda Thakkar":"hr1@maxwells.in","Harshida Pandor":"hr2@maxwells.in",
    "Patel Pritesh":"maintenance@maxwells.in","Parmar Romik":"accounts@maxwells.in",
    "Ajinkya Bapat":"purchase@maxwells.in","Mayur Dod":"sales1@maxwells.in",
    "Rajvinderkaur Hundal":"Sales2@maxwells.in","Krati Gupta":"cs@maxwells.in",
    "Rajkumar Chaudhary":"rajkumar@maxwells.in","Vinu Chavda":"vinu@maxwells.in",
    "Prabhat Singh Kumar":"ceo@maxwells.in","Pooja Lokhande":"cfo@maxwells.in",
    "Chetna Bodke":"marketing@maxwells.in","Punit Pankhaniaya":"design1@maxwells.in",
}

DEPARTMENTS = {
    "Operation":["Nishit Patel"],"QA":["Vaibhav Desai"],"QC":["Vasant Sarla"],
    "HR":["Mohit Goswami","Vrunda Thakkar","Harshida Pandor"],"Maintenance":["Patel Pritesh"],
    "Account":["Pooja Lokhande"],"Purchase":["Ajinkya Bapat"],
    "Marketing":["Mayur Dod","Rajvinderkaur Hundal"],"Design":["Punit Pankhaniaya"],"Others":[],
}
FACTORY_DEPTS = {
    "Operation":["Nishit Patel"],"QA":["Vaibhav Desai"],"QC":["Vasant Sarla"],"Maintenance":["Patel Pritesh"],
}
MANAGEMENT_LIST = ["Rajkumar Chaudhary","Vinu Chavda","Prabhat Singh Kumar","Krati Gupta","Chetna Bodke"]
PASS_COLORS = {"Factory Visit":("FFD700","000000","FACTORY"),"Staff Visit":("1565C0","FFFFFF","STAFF"),"Management":("2E7D32","FFFFFF","MANAGEMENT")}
DRINKS_MENU = ["Water","Tea","Coffee","Green Tea","Black Coffee","Juice","Other"]

LOGO_MAIN = "data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHZpZXdCb3g9IjAgMCAyNDAgNjAiPjxyZWN0IHdpZHRoPSIyNDAiIGhlaWdodD0iNjAiIGZpbGw9IiMxNTY1QzAiLz48dGV4dCB4PSIxMiIgeT0iNDMiIGZvbnQtZmFtaWx5PSJBcmlhbCIgZm9udC1zaXplPSIzNiIgZm9udC13ZWlnaHQ9IjkwMCIgZmlsbD0id2hpdGUiPk1heHdlbGw8L3RleHQ+PC9zdmc+"
DEFAULT_PHOTO = "data:image/svg+xml;base64,PHN2ZyB4bWxucz0naHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmcnIHdpZHRoPSc4MCcgaGVpZ2h0PSc4MCc+PGNpcmNsZSBjeD0nNDAnIGN5PSc0MCcgcj0nNDAnIGZpbGw9JyNlMGUwZTAnLz48dGV4dCB4PSc0MCcgeT0nNTAnIGZvbnQtZmFtaWx5PSdBcmlhbCcgZm9udC1zaXplPSczMicgZmlsbD0nIzk5OScgdGV4dC1hbmNob3I9J21pZGRsZSc+PzwvdGV4dD48L3N2Zz4="

def get_ist():
    ist = timezone(timedelta(hours=5, minutes=30))
    return datetime.now(ist).strftime("%d-%m-%Y %H:%M")

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def init_db():
    conn = sqlite3.connect(DB)
    conn.execute("""CREATE TABLE IF NOT EXISTS visitors (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, phone TEXT,
        purpose TEXT, host_name TEXT, category TEXT, department TEXT,
        person_to_meet TEXT, photo TEXT, status TEXT DEFAULT 'pending',
        created_at TEXT, checkout_at TEXT, pass_number TEXT, advance_token TEXT,
        id_front TEXT, id_back TEXT, exit_at TEXT,
        reschedule_date TEXT, reschedule_time TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS pantry_orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT, visitor_id INTEGER,
        visitor_name TEXT, person_to_meet TEXT, drink TEXT, snacks TEXT,
        quantity INTEGER, note TEXT, order_type TEXT DEFAULT 'drink',
        status TEXT DEFAULT 'pending', created_at TEXT, delivered_at TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS employee_passwords (
        email TEXT PRIMARY KEY, password_hash TEXT, is_default INTEGER DEFAULT 1
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS employee_profiles (
        email TEXT PRIMARY KEY, name TEXT, department TEXT, designation TEXT, photo TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS app_settings (key TEXT PRIMARY KEY, value TEXT)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS scheduled_meetings (
        id INTEGER PRIMARY KEY AUTOINCREMENT, host_name TEXT, visitor_name TEXT,
        visitor_phone TEXT, meeting_date TEXT, meeting_time TEXT,
        created_at TEXT, status TEXT DEFAULT 'scheduled'
    )""")
    for col in ["id_front TEXT","id_back TEXT","exit_at TEXT","reschedule_date TEXT","reschedule_time TEXT","checkout_at TEXT","advance_token TEXT"]:
        try: conn.execute("ALTER TABLE visitors ADD COLUMN {}".format(col))
        except: pass
    for col in ["snacks TEXT","order_type TEXT DEFAULT 'drink'"]:
        try: conn.execute("ALTER TABLE pantry_orders ADD COLUMN {}".format(col))
        except: pass
    conn.execute("INSERT OR IGNORE INTO app_settings (key,value) VALUES ('admin_pin','1234')")
    conn.execute("INSERT OR IGNORE INTO app_settings (key,value) VALUES ('security_pass','1234')")
    conn.commit(); conn.close()

def get_db():
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row; return conn

def get_setting(key, default=""):
    conn = get_db()
    row = conn.execute("SELECT value FROM app_settings WHERE key=?", (key,)).fetchone()
    conn.close(); return row["value"] if row else default

def set_setting(key, value):
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES (?,?)", (key,value))
    conn.commit(); conn.close()

def check_emp_password(email, password, emp_name):
    conn = get_db()
    row = conn.execute("SELECT password_hash,is_default FROM employee_passwords WHERE email=?", (email,)).fetchone()
    conn.close()
    if row: return hash_pw(password)==row["password_hash"], bool(row["is_default"])
    return password==emp_name, True

def make_qr(data):
    if not HAS_QR: return ""
    qr = qrcode.QRCode(version=1, box_size=4, border=2)
    qr.add_data(data); qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO(); img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()

def get_employee_profile(email):
    conn = get_db()
    row = conn.execute("SELECT * FROM employee_profiles WHERE email=?", (email,)).fetchone()
    conn.close(); return dict(row) if row else {}

def save_employee_profile(email, name, department, designation, photo=None):
    conn = get_db()
    if photo:
        conn.execute("INSERT OR REPLACE INTO employee_profiles (email,name,department,designation,photo) VALUES (?,?,?,?,?)",
                     (email,name,department,designation,photo))
    else:
        conn.execute("INSERT OR REPLACE INTO employee_profiles (email,name,department,designation,photo) VALUES (?,?,?,?,COALESCE((SELECT photo FROM employee_profiles WHERE email=?),NULL))",
                     (email,name,department,designation,email))
    conn.commit(); conn.close()

BEEP_JS = ("var _audioCtx=null;"
    "function _getAudioCtx(){try{if(!_audioCtx){_audioCtx=new(window.AudioContext||window.webkitAudioContext)();}"
    "if(_audioCtx.state==='suspended'){_audioCtx.resume();}return _audioCtx;}catch(e){return null;}}"
    "function _beep(n){try{var c=_getAudioCtx();if(!c)return;var f=[880,660,880,1100];"
    "for(var i=0;i<(n||4);i++){(function(x){var o=c.createOscillator(),g=c.createGain();"
    "o.connect(g);g.connect(c.destination);o.frequency.value=f[x%f.length];o.type='sine';"
    "var t=c.currentTime+x*0.22;g.gain.setValueAtTime(0,t);g.gain.linearRampToValueAtTime(0.8,t+0.05);"
    "g.gain.exponentialRampToValueAtTime(0.001,t+0.38);o.start(t);o.stop(t+0.38);})(i);}}catch(e){}}"
    "var _soundEnabled=false;"
    "function _enableSound(){if(_audioCtx)_audioCtx.resume();_soundEnabled=true;"
    "var b=document.getElementById('snd-btn');if(b)b.style.display='none';"
    "var ok=document.getElementById('snd-ok');if(ok){ok.style.display='block';setTimeout(function(){ok.style.display='none';},2000);}}"
    "setTimeout(function(){var b=document.getElementById('snd-btn');if(b)b.style.display='flex';},800);")

SOUND_WIDGET = ("<div id='snd-btn' onclick='_enableSound()' style='display:none;position:fixed;bottom:20px;right:20px;"
    "background:#1565C0;color:white;padding:11px 18px;border-radius:30px;cursor:pointer;"
    "font-weight:700;font-size:13px;box-shadow:0 4px 15px rgba(0,0,0,0.3);z-index:9999;"
    "align-items:center;gap:8px'>&#128266; Enable Sound</div>"
    "<div id='snd-ok' style='display:none;position:fixed;bottom:20px;right:20px;"
    "background:#2E7D32;color:white;padding:11px 18px;border-radius:30px;"
    "font-weight:700;font-size:13px;z-index:9999'>&#10003; Sound On!</div>")

HEADER_CSS = """.header{background:#1565C0;padding:14px 20px 10px;display:flex;flex-direction:column;align-items:center;box-shadow:0 2px 10px rgba(0,0,0,0.25)}
.hdr-logo-wrap{display:flex;justify-content:center;width:100%;margin-bottom:8px}
.hdr-logo{height:50px;object-fit:contain;filter:brightness(0) invert(1)}
.hdr-nav{display:flex;flex-wrap:wrap;justify-content:center;gap:8px}
.hdr-nav a{color:white;text-decoration:none;font-size:12px;padding:5px 12px;border:1.5px solid rgba(255,255,255,0.6);border-radius:6px;font-weight:600}
.hdr-nav a:hover{background:rgba(255,255,255,0.15)}"""

def make_header(logo, links_html):
    return ('<div class="header"><div class="hdr-logo-wrap"><img src="' + logo + '" class="hdr-logo" alt="Maxwell"></div>'
            '<div class="hdr-nav">' + links_html + '</div></div>')

# ══════════════════════════════════════════════════
# VISITOR FORM
# ══════════════════════════════════════════════════
from flask import Flask, request, jsonify, redirect, session, send_file, Response

@app.route("/")
def index():
    depts_js = json.dumps(DEPARTMENTS)
    factory_js = json.dumps(FACTORY_DEPTS)
    header = make_header(LOGO_MAIN, '<a href="/admin">Admin</a><a href="/employee-login">Employee</a><a href="/pantry-login">Pantry</a><a href="/security-login">Security</a>')
    return ("""<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Maxwell - Visitor Entry</title><style>
*{box-sizing:border-box;margin:0;padding:0}body{font-family:'Segoe UI',Arial,sans-serif;background:#f0f4f8}
""" + HEADER_CSS + """
.container{max-width:640px;margin:20px auto;padding:0 12px}
.card{background:white;border-radius:12px;padding:18px 20px;box-shadow:0 3px 10px rgba(0,0,0,0.07);margin-bottom:14px}
.section-title{font-size:14px;font-weight:700;color:#333;margin-bottom:14px}
.field-label{font-size:13px;font-weight:600;color:#555;margin-bottom:5px;display:block}.req{color:#e53935}
input[type=text],input[type=tel],textarea,select{width:100%;padding:10px 12px;border:1.5px solid #ddd;border-radius:8px;font-size:14px;background:#fafafa;font-family:inherit;outline:none}
input:focus,textarea:focus,select:focus{border-color:#1565C0;background:white}
.row2{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px}.row1{margin-bottom:12px}
.cat-chip input[type=radio]{display:none}
.cat-chip label{display:inline-block;padding:8px 18px;border-radius:25px;border:2px solid #ddd;font-size:13px;font-weight:600;cursor:pointer;color:#555;background:white;transition:0.2s}
.cat-chip input:checked+label{background:#1565C0;color:white;border-color:#1565C0}
.chip{padding:8px 15px;border-radius:20px;border:1.5px solid #ddd;font-size:13px;font-weight:600;cursor:pointer;color:#555;background:white;transition:0.2s}
.chip:hover,.chip.selected{background:#1565C0;color:white;border-color:#1565C0}
.hidden{display:none}
.btn{padding:8px 16px;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;margin:3px}
.btn-blue{background:#1565C0;color:white}.btn-green{background:#2E7D32;color:white}.btn-red{background:#C62828;color:white}.btn-purple{background:#7B1FA2;color:white}
.id-side{background:#F8F9FA;border-radius:10px;padding:14px;margin-bottom:12px;border:1.5px solid #E0E0E0}
.id-side-lbl{font-size:13px;font-weight:700;color:#333;margin-bottom:10px}
.id-icon-big{font-size:44px;color:#ddd;text-align:center;display:block;margin:4px 0}
.id-preview-img{max-width:100%;max-height:140px;border-radius:8px;border:2px solid #1565C0;display:none;margin:8px auto;display:block}
.id-btn-row{display:flex;flex-wrap:wrap;gap:6px;justify-content:center;margin-top:8px}
.submit-btn{width:100%;padding:14px;font-size:15px;border-radius:10px;background:linear-gradient(135deg,#1565C0,#1976D2);color:white;border:none;cursor:pointer;font-weight:700;letter-spacing:1px}
.alert{padding:11px 16px;border-radius:8px;margin-bottom:12px;font-size:14px;font-weight:500}
.alert-error{background:#ffebee;color:#C62828;border:1px solid #ef9a9a}
@media(max-width:540px){.row2{grid-template-columns:1fr}}
</style></head><body>""" + header + SOUND_WIDGET + """
<div class="container">
<div id="alert-box"></div>
<div id="form-section">
<div class="card"><div class="section-title">&#128203; Visitor Entry</div>
<div class="row2">
  <div><label class="field-label">Full Name <span class="req">*</span></label><input type="text" id="v-name" placeholder="Full name"></div>
  <div><label class="field-label">Phone <span class="req">*</span></label><input type="tel" id="v-phone" placeholder="Mobile"></div>
</div>
<div class="row1"><label class="field-label">Company / Organization <span class="req">*</span></label><input type="text" id="v-company" placeholder="Company name"></div>
<div class="row1"><label class="field-label">Name of Host</label><input type="text" id="v-host" placeholder="Who are you visiting?"></div>
<div class="row1"><label class="field-label">Purpose <span class="req">*</span></label><textarea id="v-purpose" rows="2" placeholder="Brief purpose"></textarea></div>
</div>
<div class="card"><div class="section-title">Visitor Category <span class="req">*</span></div>
<div style="display:flex;flex-wrap:wrap;gap:8px">
  <div class="cat-chip"><input type="radio" name="category" id="cat-f" value="Factory Visit" onchange="onCat()"><label for="cat-f">Factory Visit</label></div>
  <div class="cat-chip"><input type="radio" name="category" id="cat-s" value="Staff Visit" onchange="onCat()"><label for="cat-s">Staff Visit</label></div>
  <div class="cat-chip"><input type="radio" name="category" id="cat-m" value="Management" onchange="onCat()"><label for="cat-m">Management</label></div>
</div></div>
<div class="card hidden" id="dept-card"><div class="section-title">Department <span class="req">*</span></div><div style="display:flex;flex-wrap:wrap;gap:8px" id="dept-group"></div></div>
<div class="card hidden" id="person-card"><div class="section-title">Person to Meet <span class="req">*</span></div>
<div style="display:flex;flex-wrap:wrap;gap:8px" id="person-group"></div>
<div id="others-input" class="hidden" style="margin-top:10px"><input type="text" id="v-others" placeholder="Enter name"></div></div>
<div class="card"><div class="section-title">&#128247; Visitor Photo</div>
<div style="text-align:center;padding:8px 0">
  <span id="ph-icon" style="font-size:55px;color:#ddd;display:block">&#128247;</span>
  <img id="photo-preview" src="" style="width:100px;height:100px;border-radius:50%;object-fit:cover;border:3px solid #1565C0;display:none;margin:8px auto">
  <video id="video" autoplay playsinline style="width:100%;max-width:340px;border-radius:8px;border:2px solid #ddd;display:none"></video>
  <canvas id="canvas" style="display:none"></canvas>
  <div style="margin-top:10px">
    <button class="btn btn-blue" onclick="startCam()">&#128247; Capture</button>
    <button class="btn btn-green" id="cap-btn" onclick="capPhoto()" style="display:none">&#10003; Done</button>
    <button class="btn btn-red" id="ret-btn" onclick="retake()" style="display:none">&#8635; Retake</button>
  </div>
</div></div>
<div class="card">
<div class="section-title">&#127981; ID Card / Legal Documents</div>
<div class="id-side">
  <div class="id-side-lbl">&#9654; Front Side <span class="req">*</span></div>
  <div style="text-align:center">
    <span id="id-front-icon" class="id-icon-big">&#128195;</span>
    <img id="id-front-preview" src="" class="id-preview-img">
    <video id="id-front-video" autoplay playsinline style="width:100%;max-width:320px;border-radius:8px;border:2px solid #ddd;display:none"></video>
    <canvas id="id-front-canvas" style="display:none"></canvas>
  </div>
  <div class="id-btn-row">
    <button class="btn btn-blue" onclick="startIdCam('front')">&#128247; Camera</button>
    <button class="btn btn-green" id="id-front-cap" onclick="capId('front')" style="display:none">&#10003; Capture</button>
    <button class="btn btn-red" id="id-front-ret" onclick="retakeId('front')" style="display:none">&#8635; Retake</button>
    <label class="btn btn-purple" style="cursor:pointer">&#128444; Gallery<input type="file" accept="image/*" style="display:none" onchange="handleIdGallery('front',this)"></label>
  </div>
</div>
<div class="id-side">
  <div class="id-side-lbl">&#9654; Back Side <span style="color:#888;font-size:11px">(Optional)</span></div>
  <div style="text-align:center">
    <span id="id-back-icon" class="id-icon-big">&#128195;</span>
    <img id="id-back-preview" src="" class="id-preview-img">
    <video id="id-back-video" autoplay playsinline style="width:100%;max-width:320px;border-radius:8px;border:2px solid #ddd;display:none"></video>
    <canvas id="id-back-canvas" style="display:none"></canvas>
  </div>
  <div class="id-btn-row">
    <button class="btn btn-blue" onclick="startIdCam('back')">&#128247; Camera</button>
    <button class="btn btn-green" id="id-back-cap" onclick="capId('back')" style="display:none">&#10003; Capture</button>
    <button class="btn btn-red" id="id-back-ret" onclick="retakeId('back')" style="display:none">&#8635; Retake</button>
    <label class="btn btn-purple" style="cursor:pointer">&#128444; Gallery<input type="file" accept="image/*" style="display:none" onchange="handleIdGallery('back',this)"></label>
  </div>
</div>
</div>
<button class="submit-btn" onclick="submitForm()">SUBMIT</button>
</div>
<div id="submitted-section" class="hidden">
  <div class="card" style="text-align:center;padding:25px">
    <div style="font-size:55px">&#10003;</div><h2 style="color:#2E7D32">Submitted!</h2>
    <p style="color:#666;margin-top:8px">Awaiting approval.</p>
    <br><button class="btn btn-blue" onclick="location.reload()">+ New</button>
  </div>
</div>
</div>
<script>
var DEPTS=""" + depts_js + """;var FACTORY=""" + factory_js + """;var MGMT=""" + json.dumps(MANAGEMENT_LIST) + """;
var photoData=null,camStream=null,selDept="",selPerson="";
var idFrontData=null,idBackData=null,idCamStream={};
function onCat(){var cat=document.querySelector('input[name="category"]:checked');if(!cat)return;
  var val=cat.value;selDept='';selPerson='';
  document.getElementById('dept-card').classList.add('hidden');document.getElementById('person-card').classList.add('hidden');
  document.getElementById('dept-group').innerHTML='';document.getElementById('person-group').innerHTML='';
  if(val==='Factory Visit'){document.getElementById('dept-card').classList.remove('hidden');Object.keys(FACTORY).forEach(function(d){addChip('dept-group',d,function(){selectDept(d,val);});});}
  else if(val==='Staff Visit'){document.getElementById('dept-card').classList.remove('hidden');Object.keys(DEPTS).forEach(function(d){addChip('dept-group',d,function(){selectDept(d,val);});});}
  else if(val==='Management'){document.getElementById('person-card').classList.remove('hidden');showPersons(MGMT);}
}
function addChip(g,l,fn){var el=document.createElement('div');el.className='chip';el.textContent=l;el.onclick=fn;document.getElementById(g).appendChild(el);}
function selectDept(dept,cat){selDept=dept;selPerson='';
  document.querySelectorAll('#dept-group .chip').forEach(function(c){c.classList.remove('selected');});event.target.classList.add('selected');
  document.getElementById('person-group').innerHTML='';document.getElementById('others-input').classList.add('hidden');
  if(dept==='Others'){document.getElementById('person-card').classList.remove('hidden');document.getElementById('others-input').classList.remove('hidden');return;}
  document.getElementById('person-card').classList.remove('hidden');
  var list=(cat==='Factory Visit')?FACTORY[dept]:DEPTS[dept];if(list)showPersons(list);}
function showPersons(list){var pg=document.getElementById('person-group');pg.innerHTML='';
  list.forEach(function(p){var c=document.createElement('div');c.className='chip';c.textContent=p;
    c.onclick=function(){selPerson=p;document.querySelectorAll('#person-group .chip').forEach(function(x){x.classList.remove('selected');});c.classList.add('selected');};pg.appendChild(c);});}
async function startCam(){try{camStream=await navigator.mediaDevices.getUserMedia({video:{facingMode:'user'}});
  document.getElementById('video').srcObject=camStream;document.getElementById('video').style.display='block';
  document.getElementById('ph-icon').style.display='none';document.getElementById('photo-preview').style.display='none';
  document.getElementById('cap-btn').style.display='';document.getElementById('ret-btn').style.display='none';
}catch(e){alert('Camera denied!');}}
function capPhoto(){var v=document.getElementById('video'),c=document.getElementById('canvas');
  c.width=v.videoWidth||320;c.height=v.videoHeight||320;c.getContext('2d').drawImage(v,0,0);
  photoData=c.toDataURL('image/jpeg',0.8);document.getElementById('photo-preview').src=photoData;
  document.getElementById('photo-preview').style.display='block';document.getElementById('video').style.display='none';
  document.getElementById('cap-btn').style.display='none';document.getElementById('ret-btn').style.display='';
  if(camStream)camStream.getTracks().forEach(function(t){t.stop();});}
function retake(){photoData=null;document.getElementById('photo-preview').style.display='none';
  document.getElementById('ret-btn').style.display='none';document.getElementById('ph-icon').style.display='block';startCam();}
async function startIdCam(side){try{
  if(idCamStream[side])idCamStream[side].getTracks().forEach(function(t){t.stop();});
  idCamStream[side]=await navigator.mediaDevices.getUserMedia({video:{facingMode:'environment'}});
  var vid=document.getElementById('id-'+side+'-video');vid.srcObject=idCamStream[side];vid.style.display='block';
  document.getElementById('id-'+side+'-icon').style.display='none';document.getElementById('id-'+side+'-preview').style.display='none';
  document.getElementById('id-'+side+'-cap').style.display='';document.getElementById('id-'+side+'-ret').style.display='none';
}catch(e){alert('Camera denied!');}}
function capId(side){var v=document.getElementById('id-'+side+'-video');var c=document.getElementById('id-'+side+'-canvas');
  c.width=v.videoWidth||640;c.height=v.videoHeight||480;c.getContext('2d').drawImage(v,0,0);
  var data=c.toDataURL('image/jpeg',0.85);if(side==='front')idFrontData=data;else idBackData=data;
  document.getElementById('id-'+side+'-preview').src=data;document.getElementById('id-'+side+'-preview').style.display='block';
  document.getElementById('id-'+side+'-video').style.display='none';document.getElementById('id-'+side+'-cap').style.display='none';
  document.getElementById('id-'+side+'-ret').style.display='';document.getElementById('id-'+side+'-icon').style.display='none';
  if(idCamStream[side])idCamStream[side].getTracks().forEach(function(t){t.stop();});}
function retakeId(side){if(side==='front')idFrontData=null;else idBackData=null;
  document.getElementById('id-'+side+'-preview').style.display='none';document.getElementById('id-'+side+'-ret').style.display='none';
  document.getElementById('id-'+side+'-icon').style.display='block';startIdCam(side);}
function handleIdGallery(side,input){if(!input.files||!input.files[0])return;
  var reader=new FileReader();reader.onload=function(e){var data=e.target.result;
    if(side==='front')idFrontData=data;else idBackData=data;
    document.getElementById('id-'+side+'-preview').src=data;document.getElementById('id-'+side+'-preview').style.display='block';
    document.getElementById('id-'+side+'-icon').style.display='none';document.getElementById('id-'+side+'-ret').style.display='';};
  reader.readAsDataURL(input.files[0]);}
function showAlert(msg){document.getElementById('alert-box').innerHTML='<div class="alert alert-error">'+msg+'</div>';setTimeout(function(){document.getElementById('alert-box').innerHTML='';},5000);}
async function submitForm(){
  var name=document.getElementById('v-name').value.trim();var phone=document.getElementById('v-phone').value.trim();
  var purpose=document.getElementById('v-purpose').value.trim();var host=document.getElementById('v-host').value.trim();
  var catEl=document.querySelector('input[name="category"]:checked');var cat=catEl?catEl.value:'';
  var dept=selDept||'Management';var person=selPerson;
  if(dept==='Others')person=document.getElementById('v-others').value.trim();
  var company=document.getElementById('v-company').value.trim();
  if(!name){showAlert('Please enter visitor name!');return;}
  if(!phone){showAlert('Please enter phone!');return;}
  if(!company){showAlert('Please enter company!');return;}
  if(!purpose){showAlert('Please enter purpose!');return;}
  if(!cat){showAlert('Please select category!');return;}
  if(!person){showAlert('Please select person to meet!');return;}
  if(!idFrontData){showAlert('Please capture ID Card Front Side!');return;}
  try{var res=await fetch('/api/visitor',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({name:name,phone:phone,purpose:purpose,host_name:host,company:company,
      category:cat,department:dept,person_to_meet:person,photo:photoData,id_front:idFrontData,id_back:idBackData})});
    var data=await res.json();
    if(data.success){document.getElementById('form-section').classList.add('hidden');document.getElementById('submitted-section').classList.remove('hidden');}
    else{showAlert('Error: '+(data.error||'Unknown'));}}catch(e){showAlert('Server error!');}}
</script></body></html>""")

@app.route("/api/visitor", methods=["POST"])
def create_visitor():
    data = request.get_json(); now = get_ist(); conn = get_db(); c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM visitors"); count = c.fetchone()[0] + 1
    pass_num = "MW-" + datetime.now().strftime("%Y%m%d") + "-" + str(count).zfill(4)
    c.execute("""INSERT INTO visitors (name,phone,purpose,host_name,category,department,person_to_meet,photo,created_at,pass_number,id_front,id_back) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (data.get("name"),data.get("phone"),data.get("purpose"),data.get("host_name"),
         data.get("category"),data.get("department"),data.get("person_to_meet"),data.get("photo"),
         now,pass_num,data.get("id_front"),data.get("id_back")))
    conn.commit(); conn.close(); return jsonify({"success":True})

@app.route("/api/pending-count")
def pending_count():
    conn=get_db();row=conn.execute("SELECT COUNT(*) as cnt FROM visitors WHERE status='pending'").fetchone();conn.close()
    return jsonify({"count":row["cnt"]})

@app.route("/api/latest-pending")
def latest_pending():
    host=request.args.get("host","");conn=get_db()
    if host: row=conn.execute("SELECT id,name,person_to_meet,created_at FROM visitors WHERE status='pending' AND person_to_meet=? ORDER BY id DESC LIMIT 1",(host,)).fetchone()
    else: row=conn.execute("SELECT id,name,person_to_meet,created_at FROM visitors WHERE status='pending' ORDER BY id DESC LIMIT 1").fetchone()
    conn.close(); return jsonify({"visitor":dict(row) if row else None})

@app.route("/api/latest-visitor")
def latest_visitor():
    conn=get_db();rows=[dict(r) for r in conn.execute("SELECT id,name,phone,department,person_to_meet,status,created_at,pass_number FROM visitors ORDER BY id DESC LIMIT 20").fetchall()];conn.close()
    return jsonify({"visitors":rows})

@app.route("/api/checkout-notify")
def checkout_notify():
    conn=get_db();row=conn.execute("SELECT id,name,person_to_meet,checkout_at FROM visitors WHERE checkout_at IS NOT NULL ORDER BY id DESC LIMIT 1").fetchone();conn.close()
    return jsonify({"checkout":dict(row) if row else None})

@app.route("/api/latest-scheduled")
def latest_scheduled():
    conn=get_db();row=conn.execute("SELECT id,host_name,visitor_name,meeting_date,meeting_time,created_at FROM scheduled_meetings ORDER BY id DESC LIMIT 1").fetchone();conn.close()
    return jsonify({"meeting":dict(row) if row else None})

@app.route("/action/<int:vid>/<action>")
def do_action(vid, action):
    if action not in ("approve","reject"): return "Invalid",400
    conn=get_db();status_val="approved" if action=="approve" else "rejected"
    conn.execute("UPDATE visitors SET status=? WHERE id=?",(status_val,vid));conn.commit()
    if action=="approve":
        v=dict(conn.execute("SELECT * FROM visitors WHERE id=?",(vid,)).fetchone());conn.close()
        now=get_ist();conn2=get_db()
        conn2.execute("INSERT INTO pantry_orders (visitor_id,visitor_name,person_to_meet,drink,quantity,note,created_at,order_type) VALUES (?,?,?,?,?,?,?,?)",
                      (vid,v["name"],v["person_to_meet"],"Water",1,"Guest arrived",now,"arrival"))
        conn2.commit();conn2.close()
    else: conn.close()
    if "application/json" in request.headers.get("Accept",""):
        return jsonify({"success":True})
    return "<h2 style='text-align:center;margin-top:80px;font-family:Arial;color:#1565C0'>Visitor "+action.title()+"d!<br><a href='/admin'>Admin</a></h2>"

@app.route("/api/checkout/<int:vid>", methods=["POST"])
def checkout_visitor(vid):
    now=get_ist();conn=get_db();conn.execute("UPDATE visitors SET checkout_at=? WHERE id=?",(now,vid));conn.commit();conn.close()
    return jsonify({"success":True,"checkout_at":now})

@app.route("/api/security-exit/<int:vid>", methods=["POST"])
def security_exit(vid):
    now=get_ist();conn=get_db();conn.execute("UPDATE visitors SET exit_at=? WHERE id=?",(now,vid));conn.commit();conn.close()
    return jsonify({"success":True,"exit_at":now})

@app.route("/api/beverage", methods=["POST"])
def order_beverage():
    data=request.get_json();now=get_ist();conn=get_db()
    conn.execute("INSERT INTO pantry_orders (visitor_id,visitor_name,person_to_meet,drink,snacks,quantity,note,created_at,order_type) VALUES (?,?,?,?,?,?,?,?,?)",
                 (data.get("visitor_id"),data.get("visitor_name"),data.get("person_to_meet"),data.get("drink",""),data.get("snacks",""),data.get("quantity",1),data.get("note",""),now,"order"))
    conn.commit();conn.close();return jsonify({"success":True})

@app.route("/api/schedule-meeting", methods=["POST"])
def schedule_meeting():
    data=request.get_json();now=get_ist();conn=get_db()
    conn.execute("INSERT INTO scheduled_meetings (host_name,visitor_name,visitor_phone,meeting_date,meeting_time,created_at) VALUES (?,?,?,?,?,?)",
                 (data.get("host_name"),data.get("visitor_name"),data.get("visitor_phone"),data.get("meeting_date"),data.get("meeting_time"),now))
    conn.commit();conn.close();return jsonify({"success":True})

@app.route("/api/reschedule/<int:vid>", methods=["POST"])
def reschedule_visitor(vid):
    data=request.get_json();conn=get_db()
    conn.execute("UPDATE visitors SET status='rescheduled',reschedule_date=?,reschedule_time=? WHERE id=?",(data.get("date"),data.get("time"),vid))
    conn.commit();conn.close();return jsonify({"success":True})

@app.route("/api/pantry-pending")
def pantry_pending():
    conn=get_db();row=conn.execute("SELECT COUNT(*) as cnt FROM pantry_orders WHERE status='pending'").fetchone();conn.close()
    return jsonify({"count":row["cnt"]})

@app.route("/api/pantry-deliver/<int:oid>", methods=["POST"])
def pantry_deliver(oid):
    now=get_ist();conn=get_db();conn.execute("UPDATE pantry_orders SET status='delivered',delivered_at=? WHERE id=?",(now,oid));conn.commit();conn.close()
    return jsonify({"success":True})

@app.route("/api/admin-reset-password", methods=["POST"])
def admin_reset_password():
    if not session.get("admin_ok"): return jsonify({"error":"Not authorized"}),401
    data=request.get_json();conn=get_db()
    conn.execute("INSERT OR REPLACE INTO employee_passwords (email,password_hash,is_default) VALUES (?,?,0)",(data.get("email"),hash_pw(data.get("new_password",""))))
    conn.commit();conn.close();return jsonify({"success":True})

@app.route("/api/profile", methods=["GET"])
def get_profile():
    if not session.get("emp_email"): return jsonify({"error":"Not logged in"}),401
    email=session["emp_email"];profile=get_employee_profile(email);emp_name=session.get("emp_name","")
    dept=""
    for d,members in DEPARTMENTS.items():
        if emp_name in members: dept=d; break
    return jsonify({"name":profile.get("name",emp_name),"email":email,"department":profile.get("department",dept),"designation":profile.get("designation",""),"photo":profile.get("photo",DEFAULT_PHOTO)})

@app.route("/api/profile", methods=["POST"])
def save_profile():
    if not session.get("emp_email"): return jsonify({"error":"Not logged in"}),401
    data=request.get_json();email=session["emp_email"]
    save_employee_profile(email,data.get("name",""),data.get("department",""),data.get("designation",""),data.get("photo",None))
    return jsonify({"success":True})

@app.route("/pass/<int:vid>")
def show_pass(vid):
    conn=get_db();row=conn.execute("SELECT * FROM visitors WHERE id=?",(vid,)).fetchone();conn.close()
    if not row: return "Not found",404
    v=dict(row)
    if v["status"]!="approved":
        return "<h2 style='text-align:center;margin-top:80px;font-family:Arial;color:#C62828'>Pass not available.</h2>"
    bg,fg,label=PASS_COLORS.get(v["category"],("1565C0","FFFFFF","VISITOR"))
    photo=v.get("photo") or DEFAULT_PHOTO
    qr=make_qr("Maxwell\nName: "+v["name"]+"\nPass: "+v["pass_number"]+"\nTime: "+v["created_at"])
    qr_img='<img src="data:image/png;base64,'+qr+'" width="90">' if qr else ""
    id_html=""
    if v.get("id_front"):
        id_html="<div style='font-size:11px;color:#888;text-transform:uppercase;margin-bottom:2px'>ID (Front)</div><img src='"+v["id_front"]+"' style='max-width:120px;border-radius:5px;border:1px solid #ddd;margin-bottom:8px'>"
    return ("<!DOCTYPE html><html><head><meta charset='UTF-8'><title>Visitor Pass</title>"
            "<style>*{-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important}"
            "body{font-family:Arial;background:#f0f0f0;display:flex;justify-content:center;padding:30px;flex-direction:column;align-items:center}"
            ".pass{width:370px;border-radius:14px;overflow:hidden;box-shadow:0 8px 25px rgba(0,0,0,0.2)}"
            ".ph{padding:14px 20px;display:flex;justify-content:space-between;align-items:center;background:#"+bg+";color:#"+fg+"}"
            ".pb{background:white;padding:20px}.pf{background:#f5f5f5;text-align:center;padding:10px;font-size:11px;color:#888}"
            ".pp{width:78px;height:78px;border-radius:50%;object-fit:cover;border:3px solid white}"
            ".il{font-size:11px;color:#888;text-transform:uppercase;margin-bottom:2px}.iv{font-size:14px;font-weight:600;color:#222;margin-bottom:10px}"
            ".appr{background:#E8F5E9;color:#2E7D32;text-align:center;padding:8px;border-radius:5px;font-weight:700;margin-top:10px}"
            ".btns{margin-top:20px;text-align:center}"
            "@media print{body{background:white;padding:0}.btns{display:none}}</style></head><body>"
            "<div class='pass'><div class='ph'><span style='font-size:18px;font-weight:900'>Maxwell</span>"
            "<div style='text-align:right'><div style='font-size:20px;font-weight:900'>"+label+"</div>"
            "<div style='font-size:12px;opacity:0.8'>"+str(v["pass_number"])+"</div></div></div>"
            "<div class='pb'>"
            "<div style='display:flex;align-items:center;gap:14px;margin-bottom:18px'>"
            "<img src='"+photo+"' class='pp'>"
            "<div><div style='font-size:19px;font-weight:900;color:#1565C0'>"+str(v["name"])+"</div>"
            "<div style='font-size:12px;color:#666'>"+str(v["category"])+"</div></div></div>"
            "<hr style='border:none;border-top:1px solid #eee;margin-bottom:14px'>"
            "<div class='il'>Department</div><div class='iv'>"+str(v["department"])+"</div>"
            "<div class='il'>Person to Meet</div><div class='iv'>"+str(v["person_to_meet"])+"</div>"
            "<div class='il'>Purpose</div><div class='iv'>"+str(v["purpose"])+"</div>"
            "<div class='il'>Date & Time (IST)</div><div class='iv'>"+str(v["created_at"])+"</div>"
            +id_html+
            "<div style='text-align:center;margin-top:12px'>"+qr_img+"</div>"
            "<div class='appr'>&#10003; APPROVED</div></div>"
            "<div class='pf'>Maxwell Engineering Solutions | Visitor Management System</div></div>"
            "<div class='btns'><button onclick='window.print()' style='background:#1565C0;color:white;padding:10px 20px;border:none;border-radius:7px;cursor:pointer;font-size:14px'>Print Pass</button></div>"
            "<script>setTimeout(function(){window.print();},600);</script></body></html>")

# ══════════════════════════════════════════════════
# ADMIN
# ══════════════════════════════════════════════════
def admin_login_page(err):
    return ("<!DOCTYPE html><html><head><meta charset='UTF-8'><title>Admin Login</title>"
            "<style>body{font-family:Arial;background:#f0f4f8}.box{max-width:380px;margin:80px auto;background:white;padding:38px;border-radius:13px;box-shadow:0 5px 18px rgba(0,0,0,0.1);text-align:center}"
            ".box h2{color:#1565C0;margin-bottom:22px}input{width:100%;padding:11px;border:2px solid #e0e0e0;border-radius:7px;font-size:15px;margin-bottom:14px}"
            "button{width:100%;padding:12px;background:#1565C0;color:white;border:none;border-radius:7px;font-size:15px;font-weight:700;cursor:pointer}.err{color:red;font-size:13px;margin-top:8px}</style></head><body>"
            "<div class='box'><h2>&#128274; Admin Login</h2><form method='POST' action='/admin-login'>"
            "<input type='password' name='pin' placeholder='Enter Admin PIN'><button type='submit'>LOGIN</button></form>"
            +(("<p class='err'>"+err+"</p>") if err else "")
            +"<p style='margin-top:12px'><a href='/' style='color:#1565C0'>Back to Form</a></p></div></body></html>")

@app.route("/admin")
def admin():
    if not session.get("admin_ok"): return admin_login_page("")
    session.modified=True
    tab=request.args.get("tab","pending")
    conn=get_db()
    if tab=="all": visitors=[dict(r) for r in conn.execute("SELECT * FROM visitors ORDER BY id DESC").fetchall()]
    else: visitors=[dict(r) for r in conn.execute("SELECT * FROM visitors WHERE status=? ORDER BY id DESC",(tab,)).fetchall()]
    counts={}
    for r in conn.execute("SELECT status,COUNT(*) as cnt FROM visitors GROUP BY status").fetchall():
        counts[r["status"]]=r["cnt"]
    active_visitors=[dict(r) for r in conn.execute("SELECT * FROM visitors WHERE status='approved' AND checkout_at IS NULL ORDER BY id DESC").fetchall()]
    recent_orders=[dict(r) for r in conn.execute("SELECT * FROM pantry_orders ORDER BY id DESC LIMIT 10").fetchall()]
    scheduled=[dict(r) for r in conn.execute("SELECT * FROM scheduled_meetings ORDER BY id DESC LIMIT 15").fetchall()]
    lp=conn.execute("SELECT id FROM visitors WHERE status='pending' ORDER BY id DESC LIMIT 1").fetchone()
    lco=conn.execute("SELECT id,name,checkout_at FROM visitors WHERE checkout_at IS NOT NULL ORDER BY id DESC LIMIT 1").fetchone()
    ls=conn.execute("SELECT id FROM scheduled_meetings ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    pc=counts.get("pending",0);ac=counts.get("approved",0);rc=counts.get("rejected",0)
    lp_id=lp["id"] if lp else 0; lco_id=lco["id"] if lco else 0; ls_id=ls["id"] if ls else 0
    drinks_opts="".join('<option value="'+d+'">'+d+'</option>' for d in DRINKS_MENU)
    visitor_times={str(v["id"]):v.get("created_at","") for v in active_visitors}

    rows=""
    for v in visitors:
        bc=v["status"] if v["status"] in ("pending","approved","rejected","rescheduled") else "pending"
        photo=v["photo"] if v["photo"] else DEFAULT_PHOTO
        ab=""
        if v["status"]=="pending":
            ab+='<button class="btn ba" onclick="act('+str(v["id"])+',' + "'approve'" + ')">&#10003;</button>'
            ab+='<button class="btn br" onclick="act('+str(v["id"])+',' + "'reject'" + ')">&#10007;</button>'
        if v["status"]=="approved" and not v.get("checkout_at"):
            ab+='<button class="btn bc" onclick="chk('+str(v["id"])+')">&#128682;</button>'
        ab+='<button class="btn bp" onclick="window.open(' + "'/pass/"+str(v["id"])+"'" + ')">Pass</button>'
        co=v.get("checkout_at") or "-"; ex=v.get("exit_at") or "-"
        ec="#2E7D32" if v.get("exit_at") else ("#C62828" if v.get("checkout_at") else "#888")
        rows+=("<tr><td><img src='"+photo+"' style='width:38px;height:38px;border-radius:50%;object-fit:cover;border:2px solid #ddd'></td>"
               "<td><b>"+str(v["name"])+"</b></td><td>"+str(v["phone"])+"</td>"
               "<td>"+str(v["category"])+"</td><td>"+str(v["department"])+"</td><td>"+str(v["person_to_meet"])+"</td>"
               "<td style='font-size:11px'>"+str(v["created_at"])+"</td>"
               "<td style='font-size:11px'>"+co+"</td>"
               "<td style='font-size:11px;color:"+ec+"'>"+ex+"</td>"
               "<td><span class='badge "+bc+"'>"+v["status"].upper()+"</span></td>"
               "<td>"+ab+"</td></tr>")
    if not rows: rows='<tr><td colspan="11" style="text-align:center;padding:25px;color:#999">No records</td></tr>'

    avc=""
    for v in active_visitors:
        vid=str(v["id"])
        avc+=('<div class="avc" id="avc-'+vid+'">'
              '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">'
              '<div><b style="font-size:15px">'+v["name"]+'</b><div style="font-size:11px;color:#888">'+v["person_to_meet"]+' &bull; '+v["created_at"]+'</div></div>'
              '<button onclick="chk('+vid+')" style="background:#E53935;color:white;border:none;border-radius:8px;padding:7px 12px;font-size:12px;font-weight:700;cursor:pointer">&#128682; Checkout</button></div>'
              '<div id="ao-'+vid+'" style="display:none">'
              '<select id="adrk-'+vid+'" style="width:100%;padding:9px;border:1.5px solid #ddd;border-radius:8px;font-size:13px;margin-bottom:8px"><option value="">Select drink...</option>'+drinks_opts+'</select>'
              '<div style="display:flex;align-items:center;gap:10px;margin-bottom:8px">'
              '<button onclick="changeAQty(\''+vid+'\',-1)" style="width:32px;height:32px;border-radius:50%;border:1.5px solid #ddd;background:white;font-size:18px;cursor:pointer">-</button>'
              '<span id="aqty-'+vid+'" style="font-size:16px;font-weight:900">1</span>'
              '<button onclick="changeAQty(\''+vid+'\',1)" style="width:32px;height:32px;border-radius:50%;border:1.5px solid #ddd;background:white;font-size:18px;cursor:pointer">+</button></div>'
              '<input type="text" id="asnk-'+vid+'" placeholder="Snacks (optional)" style="width:100%;padding:9px;border:1.5px solid #ddd;border-radius:8px;font-size:13px;margin-bottom:8px">'
              '<button onclick="adminOrder(\''+vid+'\',\''+v["name"].replace("'","")+'\',\''+v["person_to_meet"].replace("'","")+'\')">&#9749; Confirm Order</button></div>'
              '<div id="at-'+vid+'" style="font-size:12px;font-weight:600;color:#F57F17;margin-top:6px;text-align:center"></div></div>')
    if not avc: avc='<div style="text-align:center;padding:20px;color:#999">No active visitors</div>'

    orows=""
    for o in recent_orders:
        sb="pending" if o["status"]=="pending" else "delivered"
        orows+=("<tr><td><b>"+str(o["visitor_name"])+"</b></td><td>"+str(o["person_to_meet"])+"</td>"
                "<td>"+(str(o.get("drink","")) or "-")+" x"+str(o.get("quantity","1"))+"</td>"
                "<td>"+(str(o.get("snacks","")) or "-")+"</td>"
                "<td style='font-size:11px'>"+str(o["created_at"])+"</td>"
                "<td><span class='badge "+sb+"'>"+o["status"].upper()+"</span></td></tr>")
    if not orows: orows='<tr><td colspan="6" style="text-align:center;padding:15px;color:#999">No orders</td></tr>'

    srows=""
    for s in scheduled:
        srows+=("<tr><td><b>"+str(s["visitor_name"])+"</b></td><td>"+str(s["visitor_phone"])+"</td>"
                "<td>"+str(s["host_name"])+"</td><td>"+str(s["meeting_date"])+" "+str(s["meeting_time"])+"</td>"
                "<td style='font-size:11px'>"+str(s["created_at"])+"</td>"
                "<td><span class='badge approved'>"+s["status"].upper()+"</span></td></tr>")
    if not srows: srows='<tr><td colspan="6" style="text-align:center;padding:15px;color:#999">No scheduled meetings</td></tr>'

    erows="".join('<tr><td><b>'+n+'</b></td><td>'+e+'</td><td><button class="btn ba" onclick="resetPwd(\''+e+'\',\''+n+'\')">&#128274; Reset</button></td></tr>' for n,e in EMPLOYEE_EMAILS.items())
    tab_a={t:"" for t in ["pending","approved","rejected","all"]}; tab_a[tab]="active"
    depts_opt="".join('<option value="'+d+'">'+d+'</option>' for d in DEPARTMENTS)
    hdr=make_header(LOGO_MAIN,'<a href="/admin/export">&#128196; Excel</a><a href="/pantry">&#9749; Pantry</a><a href="/security-login">&#128110; Security</a><a href="/admin/settings">&#9881; Settings</a><a href="/">Form</a><a href="/admin-logout">Logout</a>')

    return ("<!DOCTYPE html><html><head><meta charset='UTF-8'><title>Maxwell Admin</title>"
            "<style>*{box-sizing:border-box;margin:0;padding:0}body{font-family:Arial;background:#f0f4f8}"+HEADER_CSS
            +".container{max-width:1200px;margin:20px auto;padding:0 15px}"
            ".stats{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:20px}"
            ".stat-box{background:white;border-radius:10px;padding:20px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,0.07)}"
            ".stat-num{font-size:36px;font-weight:900}.stat-lbl{font-size:12px;color:#666;margin-top:4px}"
            ".pend{color:#F57F17}.appr-c{color:#2E7D32}.reje-c{color:#C62828}"
            ".filter-bar{background:white;padding:15px;border-radius:9px;margin-bottom:15px;display:flex;gap:10px;flex-wrap:wrap;align-items:center;box-shadow:0 2px 8px rgba(0,0,0,0.07)}"
            ".filter-bar input,.filter-bar select{padding:8px 12px;border:1.5px solid #ddd;border-radius:6px;font-size:13px}"
            ".tabs{display:flex;gap:4px;margin-bottom:15px}"
            ".tab{padding:9px 18px;border:none;border-radius:7px 7px 0 0;cursor:pointer;font-weight:600;font-size:13px;background:#ddd;color:#555}"
            ".tab.active{background:white;color:#1565C0}"
            ".card{background:white;border-radius:9px;padding:18px;box-shadow:0 2px 8px rgba(0,0,0,0.07);overflow-x:auto;margin-bottom:15px}"
            ".card h3{color:#1565C0;margin-bottom:14px;font-size:15px}"
            "table{width:100%;border-collapse:collapse;min-width:900px}.ot,.st,.et{min-width:500px}"
            "th{background:#1565C0;color:white;padding:10px;text-align:left;font-size:12px}"
            "td{padding:9px;border-bottom:1px solid #eee;font-size:13px;vertical-align:middle}"
            "tr:hover td{background:#f9f9f9}"
            ".badge{display:inline-block;padding:3px 9px;border-radius:11px;font-size:11px;font-weight:700}"
            ".badge.pending{background:#FFF8E1;color:#F57F17}.badge.approved{background:#E8F5E9;color:#2E7D32}"
            ".badge.rejected{background:#FFEBEE;color:#C62828}.badge.delivered{background:#E8F5E9;color:#2E7D32}"
            ".badge.rescheduled{background:#E3F2FD;color:#1565C0}"
            ".btn{padding:5px 10px;border:none;border-radius:5px;cursor:pointer;font-size:11px;font-weight:600;margin:2px}"
            ".ba{background:#2E7D32;color:white}.br{background:#C62828;color:white}.bp{background:#1565C0;color:white}.bc{background:#FF6F00;color:white}"
            ".avc{background:#F8FAFE;border-radius:12px;padding:13px;border:1.5px solid #E3F2FD;margin-bottom:10px}"
            ".avc button[onclick*='adminOrder']{width:100%;padding:11px;background:#1565C0;color:white;border:none;border-radius:8px;font-size:14px;font-weight:700;cursor:pointer}"
            ".nb{background:#FFF8E1;border:2px solid #F57F17;border-radius:9px;padding:12px 18px;margin-bottom:15px;display:none;font-weight:700;color:#E65100;font-size:14px}"
            "</style></head><body>"
            +hdr+SOUND_WIDGET+
            "<div class='container'><div class='nb' id='nb'></div>"
            "<div class='stats'>"
            "<div class='stat-box'><div class='stat-num pend'>"+str(pc)+"</div><div class='stat-lbl'>&#9203; Pending</div></div>"
            "<div class='stat-box'><div class='stat-num appr-c'>"+str(ac)+"</div><div class='stat-lbl'>&#10003; Approved</div></div>"
            "<div class='stat-box'><div class='stat-num reje-c'>"+str(rc)+"</div><div class='stat-lbl'>&#10007; Rejected</div></div>"
            "</div>"
            "<div class='filter-bar'><input type='date' id='f-from'><input type='date' id='f-to'>"
            "<select id='f-dept'><option value=''>All Departments</option>"+depts_opt+"</select>"
            "<select id='f-status'><option value=''>All Status</option><option>pending</option><option>approved</option><option>rejected</option></select>"
            "<button onclick='doExport()' style='background:#2E7D32;color:white;padding:8px 16px;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600'>&#128196; Export</button></div>"
            "<div class='tabs'>"
            "<button class='tab "+tab_a["pending"]+"' onclick='location.href=\"?tab=pending\"'>Pending ("+str(pc)+")</button>"
            "<button class='tab "+tab_a["approved"]+"' onclick='location.href=\"?tab=approved\"'>Approved</button>"
            "<button class='tab "+tab_a["rejected"]+"' onclick='location.href=\"?tab=rejected\"'>Rejected</button>"
            "<button class='tab "+tab_a["all"]+"' onclick='location.href=\"?tab=all\"'>All</button></div>"
            "<div class='card'><table>"
            "<tr><th>Photo</th><th>Name</th><th>Phone</th><th>Category</th><th>Dept</th><th>Meeting</th><th>In Time</th><th>Out</th><th>Exit</th><th>Status</th><th>Action</th></tr>"
            +rows+"</table></div>"
            "<div class='card'><h3>&#9749; Active Visitors – Drink Orders</h3><div id='avc-wrap'>"+avc+"</div></div>"
            "<div class='card'><h3>&#9749; Recent Pantry Orders</h3><table class='ot'>"
            "<tr><th>Guest</th><th>Host</th><th>Drink</th><th>Snacks</th><th>Time</th><th>Status</th></tr>"+orows+"</table></div>"
            "<div class='card'><h3>&#128197; Scheduled Meetings</h3><table class='st'>"
            "<tr><th>Visitor</th><th>Phone</th><th>Host</th><th>Date &amp; Time</th><th>Created</th><th>Status</th></tr>"+srows+"</table></div>"
            "<div class='card'><h3>&#128274; Employee Password Reset</h3>"
            "<div id='rst-msg' style='display:none;background:#E8F5E9;color:#2E7D32;padding:10px;border-radius:7px;margin-bottom:10px;font-weight:600'></div>"
            "<table class='et'><tr><th>Employee</th><th>Email</th><th>Action</th></tr>"+erows+"</table></div>"
            "</div>"
            "<script>"+BEEP_JS
            +"var _ol="+str(lp_id)+",_oc="+str(lco_id)+",_sl="+str(ls_id)+",_aq={},_fw=true;"
            "var _vt="+json.dumps(visitor_times)+";"
            "function doExport(){var fr=document.getElementById('f-from').value,to=document.getElementById('f-to').value,dept=document.getElementById('f-dept').value,st=document.getElementById('f-status').value;window.open('/admin/export?from='+encodeURIComponent(fr)+'&to='+encodeURIComponent(to)+'&dept='+encodeURIComponent(dept)+'&status='+encodeURIComponent(st));}"
            "async function act(id,action){if(!confirm(action+' this visitor?'))return;await fetch('/action/'+id+'/'+action,{headers:{'Accept':'application/json'}});if(action==='approve')window.open('/pass/'+id);location.reload();}"
            "async function chk(id){if(!confirm('Checkout?'))return;await fetch('/api/checkout/'+id,{method:'POST'});location.reload();}"
            "function changeAQty(v,d){if(!_aq[v])_aq[v]=1;_aq[v]=Math.max(1,_aq[v]+d);document.getElementById('aqty-'+v).textContent=_aq[v];}"
            "async function adminOrder(vid,vname,person){var drink=document.getElementById('adrk-'+vid).value;var qty=_aq[vid]||1;var snacks=document.getElementById('asnk-'+vid).value;"
            "if(!drink&&!snacks){alert('Select drink or snacks!');return;}"
            "await fetch('/api/beverage',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({visitor_id:vid,visitor_name:vname,person_to_meet:person,drink:drink,quantity:qty,snacks:snacks})});"
            "showNB('&#10003; Order sent!',5000);}"
            "function parseIST(s){if(!s)return new Date();var p=s.split(' ');var dp=p[0].split('-');var tp=(p[1]||'0:0').split(':');return new Date(parseInt(dp[2]),parseInt(dp[1])-1,parseInt(dp[0]),parseInt(tp[0]),parseInt(tp[1]));}"
            "function checkAVO(){var now=new Date();Object.keys(_vt).forEach(function(vid){var ot=parseIST(_vt[vid]);var diff=Math.floor((now-ot)/60000);var ao=document.getElementById('ao-'+vid);var at=document.getElementById('at-'+vid);if(ao){if(diff>=7){ao.style.display='block';if(at)at.textContent='';}else{ao.style.display='none';if(at)at.textContent='Order in '+(7-diff)+'m';}}});}"
            "async function resetPwd(email,name){var np=prompt('New password for '+name+':');if(!np||np.length<4){alert('Min 4 chars!');return;}"
            "var r=await fetch('/api/admin-reset-password',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({email:email,new_password:np})});"
            "var d=await r.json();if(d.success){var m=document.getElementById('rst-msg');m.textContent='&#10003; Password reset for '+name+'!';m.style.display='block';setTimeout(function(){m.style.display='none';},4000);}}"
            "function showNB(msg,dur){var b=document.getElementById('nb');b.innerHTML=msg;b.style.display='block';setTimeout(function(){b.style.display='none';},dur||8000);}"
            "async function checkNew(){try{"
            "var r0=await fetch('/api/latest-pending');var d0=await r0.json();"
            "if(_fw&&d0.visitor&&d0.visitor.id>_ol){_beep(4);showNB('&#128276; New visitor: '+d0.visitor.name,10000);}"
            "if(d0.visitor)_ol=d0.visitor.id;_fw=true;"
            "var r1=await fetch('/api/pending-count');var d1=await r1.json();"
            "document.title=d1.count>0?'('+d1.count+') NEW! - Admin':'Maxwell Admin';"
            "var r2=await fetch('/api/checkout-notify');var d2=await r2.json();"
            "if(d2.checkout&&d2.checkout.id>_oc){_beep(3);showNB('&#128682; Checkout: '+d2.checkout.name,10000);_oc=d2.checkout.id;}"
            "var r3=await fetch('/api/latest-scheduled');var d3=await r3.json();"
            "if(d3.meeting&&d3.meeting.id>_sl){_beep(2);showNB('&#128197; Meeting scheduled: '+d3.meeting.visitor_name+' with '+d3.meeting.host_name,12000);_sl=d3.meeting.id;}"
            "}catch(e){}}"
            "if(Notification.permission==='default')Notification.requestPermission();"
            "setInterval(checkNew,8000);checkNew();setInterval(checkAVO,15000);checkAVO();"
            "</script></body></html>")

@app.route("/admin-login", methods=["POST"])
def admin_login():
    pin=request.form.get("pin","")
    if pin==get_setting("admin_pin","1234"):
        session["admin_ok"]=True; return redirect("/admin")
    return admin_login_page("Wrong PIN!")

@app.route("/admin-logout")
def admin_logout():
    session.clear(); return redirect("/admin")

@app.route("/admin/settings", methods=["GET","POST"])
def admin_settings():
    if not session.get("admin_ok"): return redirect("/admin")
    msg=""
    if request.method=="POST":
        action=request.form.get("action","")
        if action=="admin_pin":
            np=request.form.get("new_pin","").strip()
            if len(np)>=4: set_setting("admin_pin",np); msg="Admin PIN updated!"
            else: msg="PIN must be at least 4 digits!"
        elif action=="security_pass":
            np=request.form.get("new_pass","").strip()
            if len(np)>=4: set_setting("security_pass",np); msg="Security password updated!"
            else: msg="Password must be at least 4 characters!"
    hdr=make_header(LOGO_MAIN,'<a href="/admin">&#8592; Admin</a>')
    return ("<!DOCTYPE html><html><head><meta charset='UTF-8'><title>Settings</title>"
            "<style>body{font-family:Arial;background:#f0f4f8}"+HEADER_CSS
            +".container{max-width:600px;margin:30px auto;padding:0 15px}"
            ".card{background:white;border-radius:9px;padding:22px;box-shadow:0 2px 8px rgba(0,0,0,0.07);margin-bottom:18px}"
            ".card h3{color:#1565C0;margin-bottom:16px}label{display:block;font-size:13px;font-weight:600;color:#444;margin-bottom:5px;margin-top:12px}"
            "input{width:100%;padding:10px;border:1.5px solid #ddd;border-radius:7px;font-size:14px}input:focus{outline:none;border-color:#1565C0}"
            "button{padding:10px 22px;background:#1565C0;color:white;border:none;border-radius:7px;font-size:14px;font-weight:700;cursor:pointer;margin-top:12px}"
            ".msg{background:#E8F5E9;color:#2E7D32;padding:10px;border-radius:7px;margin-bottom:15px;font-weight:600}"
            "</style></head><body>"+hdr+
            "<div class='container'>"+(("<div class='msg'>&#10003; "+msg+"</div>") if msg else "")
            +"<div class='card'><h3>&#128274; Change Admin PIN</h3>"
            "<form method='POST'><input type='hidden' name='action' value='admin_pin'>"
            "<label>New PIN</label><input type='password' name='new_pin' placeholder='Min 4 digits' required>"
            "<button type='submit'>Update PIN</button></form></div>"
            "<div class='card'><h3>&#128110; Change Security Password</h3>"
            "<form method='POST'><input type='hidden' name='action' value='security_pass'>"
            "<label>New Password</label><input type='password' name='new_pass' placeholder='Min 4 chars' required>"
            "<button type='submit'>Update</button></form></div></div></body></html>")

@app.route("/admin/export")
def export_excel():
    if not session.get("admin_ok"): return redirect("/admin")
    if not HAS_EXCEL: return "openpyxl not installed",400
    date_from=request.args.get("from",""); date_to=request.args.get("to","")
    dept=request.args.get("dept",""); status=request.args.get("status","")
    conn=get_db(); query="SELECT * FROM visitors WHERE 1=1"; params=[]
    if date_from: query+=" AND created_at >= ?"; params.append(date_from)
    if date_to: query+=" AND created_at <= ?"; params.append(date_to+" 23:59")
    if dept: query+=" AND department=?"; params.append(dept)
    if status: query+=" AND status=?"; params.append(status)
    query+=" ORDER BY id DESC"
    visitors=[dict(r) for r in conn.execute(query,params).fetchall()]; conn.close()
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Visitor Report"
    hf=PatternFill("solid",fgColor="1565C0"); hfnt=Font(color="FFFFFF",bold=True,size=12)
    headers=["#","Name","Phone","Person","Dept","Category","Purpose","Status","Pass","In Time","Out Time","Exit Time"]
    for col,h in enumerate(headers,1):
        cell=ws.cell(row=1,column=col,value=h); cell.fill=hf; cell.font=hfnt; cell.alignment=Alignment(horizontal="center")
    ws.row_dimensions[1].height=25
    for i,w in enumerate([5,20,14,20,14,14,24,12,18,18,18,18],1):
        ws.column_dimensions[chr(64+i)].width=w
    for row,v in enumerate(visitors,2):
        ws.cell(row=row,column=1,value=row-1)
        for ci,key in enumerate(["name","phone","person_to_meet","department","category","purpose","status","pass_number","created_at","checkout_at","exit_at"],2):
            ws.cell(row=row,column=ci,value=v.get(key,"") or "-")
        fill=PatternFill("solid",fgColor="E8F5E9" if v.get("status")=="approved" else "FFEBEE" if v.get("status")=="rejected" else "FFF8E1")
        for ci in range(2,13): ws.cell(row=row,column=ci).fill=fill
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    filename="Maxwell_Visitors_"+datetime.now().strftime("%d%m%Y")+".xlsx"
    return send_file(buf,as_attachment=True,download_name=filename,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══════════════════════════════════════════════════
# PANTRY
# ══════════════════════════════════════════════════
@app.route("/pantry-login", methods=["GET","POST"])
def pantry_login():
    err=""
    if request.method=="POST":
        email=request.form.get("email","").strip(); pwd=request.form.get("password","").strip()
        if email==PANTRY_EMAIL and pwd==PANTRY_PASSWORD:
            session["pantry_ok"]=True; return redirect("/pantry")
        err="Wrong credentials!"
    return ("<!DOCTYPE html><html><head><meta charset='UTF-8'><title>Pantry Login</title>"
            "<style>body{font-family:Arial;background:#f0f4f8}.box{max-width:400px;margin:80px auto;background:white;padding:38px;border-radius:13px;box-shadow:0 5px 18px rgba(0,0,0,0.1)}"
            ".box h2{color:#1565C0;text-align:center;margin-bottom:22px}label{display:block;font-size:13px;font-weight:600;color:#444;margin-bottom:5px;margin-top:14px}"
            "input{width:100%;padding:11px;border:2px solid #e0e0e0;border-radius:7px;font-size:14px}input:focus{outline:none;border-color:#1565C0}"
            "button{width:100%;margin-top:18px;padding:12px;background:#1565C0;color:white;border:none;border-radius:7px;font-size:15px;font-weight:700;cursor:pointer}"
            ".err{color:red;font-size:13px;margin-top:8px;text-align:center}"
            "</style></head><body><div class='box'><h2>&#9749; Pantry Login</h2>"
            "<form method='POST'><label>Email</label><input type='email' name='email' value='maxwellvisitor05@gmail.com'>"
            "<label>Password</label><input type='password' name='password'><button type='submit'>LOGIN</button></form>"
            +(("<p class='err'>"+err+"</p>") if err else "")+"</div></body></html>")

@app.route("/pantry")
def pantry():
    if not session.get("pantry_ok") and not session.get("admin_ok"): return redirect("/pantry-login")
    conn=get_db()
    orders=[dict(r) for r in conn.execute("SELECT * FROM pantry_orders ORDER BY id DESC LIMIT 60").fetchall()]
    conn.close()
    order_times={str(o["id"]):o["created_at"] for o in orders if o["status"]=="pending"}
    rows=""
    for o in orders:
        sb="pending" if o["status"]=="pending" else "delivered"
        ab=timer_html=""
        if o["status"]=="pending":
            ab='<button class="btn ba" onclick="delv('+str(o["id"])+')">&#10003; Delivered</button>'
            timer_html='<div id="tmr-'+str(o["id"])+'" style="font-size:12px;font-weight:600;margin-top:4px"></div>'
        dk=str(o.get("drink","") or "-"); qty=str(o.get("quantity","1")); snk=str(o.get("snacks","")) if o.get("snacks") else "-"
        nt=str(o.get("note","") or "-")
        drink_guj={"Water":"પાણી","Tea":"ચા","Coffee":"કોફી","Green Tea":"ગ્રીન ટી","Black Coffee":"બ્લેક કોફી","Juice":"જ્યૂસ","Other":"ઓર્ડર"}
        qty_guj={"1":"એક","2":"બે","3":"ત્રણ","4":"ચાર","5":"પાંચ","6":"છ","7":"સાત","8":"આઠ","9":"નવ","10":"દસ"}
        gd=drink_guj.get(o.get("drink",""),o.get("drink",""))
        gq=qty_guj.get(str(o.get("quantity",1)),str(o.get("quantity",1)))
        guj_txt=gq+" "+gd
        if o.get("snacks"): guj_txt+=" ane "+str(o.get("snacks",""))
        guj_disp=('<div style="background:#EDE7F6;border-radius:8px;padding:8px 12px;margin-top:6px;font-size:13px;color:#4A148C;font-weight:600">&#127370; '+guj_txt+'</div>' if guj_txt else "")
        spk_data=(str(o.get("quantity",1))+" cup "+str(o.get("drink","")) if o.get("drink") else "")
        if o.get("snacks"): spk_data+=(" ane " if spk_data else "")+str(o.get("snacks",""))
        spk_btn=('<button class="btn" data-order="'+spk_data.replace('"',"&quot;")+'" data-guest="'+str(o.get("visitor_name","")).replace('"',"&quot;")+'" onclick="speakOrder(this)" style="background:#9C27B0;color:white;margin-top:4px">&#128266; બોલો</button>' if spk_data else "")
        rows+=("<tr id='row-"+str(o["id"])+"'><td><b>"+str(o["visitor_name"])+"</b></td><td>"+str(o["person_to_meet"])+"</td>"
               "<td>"+dk+(" x"+qty if dk!="-" else "")+"</td><td>"+snk+"</td><td>"+nt+"</td>"
               "<td style='font-size:11px'>"+str(o["created_at"])+"</td>"
               "<td><span class='badge "+sb+"'>"+o["status"].upper()+"</span>"+timer_html+guj_disp+spk_btn+"</td>"
               "<td>"+ab+"</td></tr>")
    if not rows: rows='<tr><td colspan="8" style="text-align:center;padding:25px;color:#999">No orders</td></tr>'
    hdr=make_header(LOGO_MAIN,'<a href="/admin">Admin</a><a href="/pantry-logout">Logout</a>')
    return ("<!DOCTYPE html><html><head><meta charset='UTF-8'><title>Pantry</title>"
            "<style>*{box-sizing:border-box;margin:0;padding:0}body{font-family:Arial;background:#f0f4f8}"+HEADER_CSS
            +".container{max-width:1050px;margin:20px auto;padding:0 15px}"
            ".card{background:white;border-radius:9px;padding:18px;box-shadow:0 2px 8px rgba(0,0,0,0.07);overflow-x:auto;margin-bottom:15px}"
            ".card h3{color:#1565C0;margin-bottom:14px}"
            ".nb{background:#FFF8E1;border:2px solid #F57F17;border-radius:9px;padding:15px;margin-bottom:15px;text-align:center;font-weight:700;color:#E65100;font-size:16px;display:none}"
            "table{width:100%;border-collapse:collapse;min-width:720px}"
            "th{background:#1565C0;color:white;padding:10px;font-size:12px;text-align:left}"
            "td{padding:10px;border-bottom:1px solid #eee;font-size:13px;vertical-align:middle}"
            ".badge{display:inline-block;padding:3px 9px;border-radius:11px;font-size:11px;font-weight:700}"
            ".badge.pending{background:#FFF8E1;color:#F57F17}.badge.delivered{background:#E8F5E9;color:#2E7D32}"
            ".btn{padding:6px 12px;border:none;border-radius:5px;cursor:pointer;font-size:12px;font-weight:600;margin:2px}.ba{background:#2E7D32;color:white}"
            ".timer-red{color:#C62828!important;font-weight:900!important}"
            "</style></head><body>"+hdr+SOUND_WIDGET+
            "<div class='container'><div class='nb' id='nb'>&#128276; NEW ORDER!</div>"
            "<div class='card'><h3>&#9749; Pantry Orders</h3>"
            "<table><tr><th>Guest</th><th>Host</th><th>Drink</th><th>Snacks</th><th>Note</th><th>Time</th><th>Status</th><th>Action</th></tr>"
            +rows+"</table></div></div>"
            "<script>"+BEEP_JS+"var _pc=0;var _dt="+json.dumps(order_times)+";"
            "async function delv(id){var r=await fetch('/api/pantry-deliver/'+id,{method:'POST'});var d=await r.json();if(d.success)location.reload();}"
            "function parseIST(s){var p=s.split(' ');var dp=p[0].split('-');var tp=p[1].split(':');return new Date(parseInt(dp[2]),parseInt(dp[1])-1,parseInt(dp[0]),parseInt(tp[0]),parseInt(tp[1]));}"
            "function updateTimers(){var now=new Date();Object.keys(_dt).forEach(function(id){"
            "var el=document.getElementById('tmr-'+id);if(!el)return;"
            "var diff=Math.floor((now-parseIST(_dt[id]))/60000);var rem=10-diff;"
            "if(rem<=0){el.innerHTML='&#9888; OVERDUE '+Math.abs(rem)+'m';el.className='timer-red';}"
            "else if(rem<=3){el.innerHTML='&#9888; '+rem+'m left';el.style.color='#C62828';}"
            "else{el.innerHTML='&#9200; '+rem+'m left';el.style.color=rem<=7?'#F57F17':'#2E7D32';}});}"
            "function speakOrder(btn){if(!window.speechSynthesis){alert('Speech not supported');return;}"
            "var order=btn.getAttribute('data-order');var guest=btn.getAttribute('data-guest');"
            "window.speechSynthesis.cancel();"
            "var dm={'Water':'પાણી','Tea':'ચા','Coffee':'કોફી','Green Tea':'ગ્રીન ટી','Black Coffee':'બ્લેક કોફી','Juice':'જ્યૂસ'};"
            "var qm={'1':'એક','2':'બે','3':'ત્રણ','4':'ચાર','5':'પાંચ','6':'છ','7':'સાત','8':'આઠ','9':'નવ','10':'દસ'};"
            "var parts=order.split(' cup ');var qty=parts[0];var rest=parts.length>1?parts[1]:order;"
            "var dp=rest.split(' ane ');var drink=dp[0];var snacks=dp.length>1?dp[1]:'';"
            "var text=guest+' mates '+( qm[qty]||qty)+' '+(dm[drink]||drink)+' lai jao';"
            "if(snacks&&snacks.indexOf('note:')<0)text+=' ane '+snacks+' lai jao';"
            "var u=new SpeechSynthesisUtterance(text);u.lang='gu-IN';u.rate=0.85;"
            "var voices=window.speechSynthesis.getVoices();"
            "var gv=voices.find(function(v){return v.lang==='gu-IN'||v.lang.startsWith('gu');});"
            "if(gv)u.voice=gv;window.speechSynthesis.speak(u);}"
            "async function checkOrders(){try{var r=await fetch('/api/pantry-pending');var d=await r.json();"
            "if(_pc>0&&d.count>_pc){_enableSound();_beep(5);if(navigator.vibrate)navigator.vibrate([300,100,300,100,300]);"
            "document.getElementById('nb').style.display='block';setTimeout(function(){document.getElementById('nb').style.display='none';},10000);"
            "location.reload();}_pc=d.count;document.title=d.count>0?'('+d.count+') NEW! - Pantry':'Pantry';"
            "}catch(e){}}"
            "setInterval(checkOrders,6000);checkOrders();setInterval(updateTimers,15000);updateTimers();"
            "window.speechSynthesis.onvoiceschanged=function(){window.speechSynthesis.getVoices();};"
            "</script></body></html>")

@app.route("/pantry-logout")
def pantry_logout():
    session.pop("pantry_ok",None); return redirect("/pantry-login")

# ══════════════════════════════════════════════════
# EMPLOYEE LOGIN + DASHBOARD
# ══════════════════════════════════════════════════
@app.route("/employee-login", methods=["GET","POST"])
def employee_login():
    err=""
    if request.method=="POST":
        email=request.form.get("email","").lower().strip(); password=request.form.get("password","").strip()
        emp_map={v.lower():k for k,v in EMPLOYEE_EMAILS.items()}
        if email in emp_map:
            emp_name=emp_map[email]; pw_ok,is_default=check_emp_password(email,password,emp_name)
            if pw_ok:
                session.permanent=True; session["emp_name"]=emp_name; session["emp_email"]=email
                if is_default: session["force_pw_change"]=True
                return redirect("/employee-dashboard")
            else: err="Wrong password!"
        else: err="Email not found."
    return ("<!DOCTYPE html><html><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width,initial-scale=1'>"
            "<title>Maxwell - Login</title>"
            "<style>*{box-sizing:border-box;margin:0;padding:0}body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;min-height:100vh;"
            "background:linear-gradient(160deg,#1565C0,#0D47A1);display:flex;flex-direction:column;align-items:center;justify-content:center;padding:20px}"
            ".la{margin-bottom:24px;text-align:center}.la img{height:60px;filter:brightness(0) invert(1)}.la p{color:rgba(255,255,255,0.75);font-size:13px;margin-top:6px}"
            ".box{background:white;border-radius:24px;padding:32px 28px;width:100%;max-width:390px;box-shadow:0 25px 50px rgba(0,0,0,0.25)}"
            ".box h2{color:#1565C0;text-align:center;font-size:22px;font-weight:800;margin-bottom:22px}"
            "label{display:block;font-size:13px;font-weight:600;color:#555;margin-bottom:6px;margin-top:16px}"
            ".pw-wrap{position:relative}.pw-wrap input{padding-right:44px}"
            ".eye-btn{position:absolute;right:13px;top:50%;transform:translateY(-50%);background:none;border:none;cursor:pointer;font-size:18px;color:#bbb}"
            "input{width:100%;padding:13px 16px;border:2px solid #eee;border-radius:12px;font-size:15px;outline:none;background:#fafafa;transition:border 0.2s}"
            "input:focus{border-color:#1565C0;background:white}"
            "button.sbtn{width:100%;margin-top:22px;padding:15px;background:linear-gradient(135deg,#1565C0,#1976D2);color:white;border:none;border-radius:12px;font-size:16px;font-weight:700;cursor:pointer}"
            ".err{color:#e53935;font-size:13px;margin-top:10px;text-align:center;background:#ffebee;padding:10px;border-radius:10px;font-weight:500}"
            "a{display:block;text-align:center;margin-top:16px;color:rgba(255,255,255,0.8);font-size:13px;text-decoration:none}"
            "</style></head><body>"
            "<div class='la'><img src='"+LOGO_MAIN+"' alt='Maxwell'><p>Maxwell Engineering Solutions</p></div>"
            "<div class='box'><h2>Welcome Back!</h2><form method='POST'>"
            "<label>Company Email</label><input type='email' name='email' placeholder='yourname@maxwells.in' required>"
            "<label>Password</label><div class='pw-wrap'><input type='password' name='password' id='pw' placeholder='Enter password' required>"
            "<button type='button' class='eye-btn' onclick='var f=document.getElementById(\"pw\");f.type=f.type===\"password\"?\"text\":\"password\"'>&#128065;</button></div>"
            "<button type='submit' class='sbtn'>LOGIN</button></form>"
            +(("<p class='err'>"+err+"</p>") if err else "")
            +"</div><a href='/'>&#8592; Back to Visitor Form</a></body></html>")

@app.route("/employee-dashboard")
def employee_dashboard():
    if not session.get("emp_name"): return redirect("/employee-login")
    session.permanent=True
    name=session["emp_name"]; email=session.get("emp_email","")
    profile=get_employee_profile(email)
    emp_photo=profile.get("photo",DEFAULT_PHOTO) or DEFAULT_PHOTO
    emp_dept=profile.get("department",""); emp_desig=profile.get("designation","")
    if not emp_dept:
        for d,members in DEPARTMENTS.items():
            if name in members: emp_dept=d; break
    conn=get_db()
    visitors=[dict(r) for r in conn.execute("SELECT * FROM visitors WHERE person_to_meet=? AND status='approved' AND checkout_at IS NULL ORDER BY id DESC",(name,)).fetchall()]
    all_visitors=[dict(r) for r in conn.execute("SELECT * FROM visitors WHERE person_to_meet=? ORDER BY id DESC LIMIT 10",(name,)).fetchall()]
    pending_visitors=[dict(r) for r in conn.execute("SELECT * FROM visitors WHERE person_to_meet=? AND status='pending' ORDER BY id DESC",(name,)).fetchall()]
    lhp=conn.execute("SELECT id FROM visitors WHERE person_to_meet=? AND status='pending' ORDER BY id DESC LIMIT 1",(name,)).fetchone()
    today=get_ist()[:10]
    today_count=conn.execute("SELECT COUNT(*) as cnt FROM visitors WHERE person_to_meet=? AND created_at LIKE ?",(name,today+"%")).fetchone()["cnt"]
    checked_out=conn.execute("SELECT COUNT(*) as cnt FROM visitors WHERE person_to_meet=? AND checkout_at IS NOT NULL AND created_at LIKE ?",(name,today+"%")).fetchone()["cnt"]
    in_house=conn.execute("SELECT COUNT(*) as cnt FROM visitors WHERE person_to_meet=? AND status='approved' AND checkout_at IS NULL",(name,)).fetchone()["cnt"]
    orders_today=conn.execute("SELECT COUNT(*) as cnt FROM pantry_orders WHERE person_to_meet=? AND created_at LIKE ?",(name,today+"%")).fetchone()["cnt"]
    conn.close()
    lhp_id=lhp["id"] if lhp else 0
    force_change=session.get("force_pw_change",False)
    drinks_opts="".join('<option value="'+d+'">'+d+'</option>' for d in DRINKS_MENU)
    visitor_times={str(v["id"]):v.get("created_at","") for v in visitors}
    pc=len(pending_visitors)

    active_cards=""
    for v in visitors:
        vid=str(v["id"])
        active_cards+=('<div class="avc" id="avc-'+vid+'"><div class="av-top">'
            '<div class="av-left"><div class="av-name">'+v["name"]+'</div>'
            '<div class="av-time">'+v["created_at"]+'</div>'
            '<div class="av-purpose">'+v["purpose"][:40]+'</div></div>'
            '<button class="co-btn" onclick="checkout('+vid+')">&#128682; Checkout</button></div>'
            '<div class="ord-sec" id="ord-'+vid+'" style="display:none">'
            '<div class="ob"><div class="ob-title">&#9749; Order Drink</div>'
            '<div class="drink-row">'
            '<select id="drk-'+vid+'" class="drink-sel"><option value="">Select drink...</option>'+drinks_opts+'</select>'
            '<div class="qty-wrap"><button class="qty-btn" onclick="changeQty(\''+vid+'\',-1)">-</button>'
            '<span id="qty-'+vid+'" class="qty-num">1</span>'
            '<button class="qty-btn" onclick="changeQty(\''+vid+'\',1)">+</button></div></div></div>'
            '<div class="ob" style="margin-top:10px"><div class="ob-title">&#127839; Snacks</div>'
            '<input type="text" id="snk-'+vid+'" placeholder="e.g. Biscuits..." class="si"></div>'
            '<div class="ob" style="margin-top:10px"><div class="ob-title">&#128221; Note</div>'
            '<input type="text" id="nte-'+vid+'" placeholder="e.g. 2 with sugar..." class="si"></div>'
            '<button class="co-btn2" onclick="confirmOrder(\''+vid+'\',\''+v["name"].replace("'","")+"','"+name.replace("'","")+"'" + ')">&#10003; Confirm Order</button></div>'
            '<div id="tl-'+vid+'" class="tl"></div></div>')
    if not active_cards:
        active_cards='<div style="text-align:center;padding:28px;color:#999"><div style="font-size:36px;margin-bottom:8px">&#128100;</div>No active visitors</div>'

    # Pending cards with Reschedule button
    pending_cards=""
    for v in pending_visitors:
        vid=str(v["id"]); photo=v.get("photo") or DEFAULT_PHOTO
        ph=str(v.get("phone","")).replace("'",""); vn=str(v["name"]).replace("'","")
        pending_cards+=('<div class="pi" id="pi-'+vid+'">'
            '<img src="'+photo+'" class="pi-photo">'
            '<div class="pi-info"><div class="pi-name">'+v["name"]+'</div>'
            '<div class="pi-sub">'+v["department"]+'</div><div class="pi-t">'+v["created_at"]+'</div></div>'
            '<div class="pi-acts">'
            '<button class="pa-ok" onclick="act('+vid+',' + "'approve'" + ')" title="Approve">&#10003;</button>'
            '<button class="pa-no" onclick="act('+vid+',' + "'reject'" + ')" title="Reject">&#10007;</button>'
            '<button class="pa-rs" onclick="toggleRs(\''+vid+'\')" title="Reschedule">&#128197;</button>'
            '</div></div>'
            '<div id="rs-'+vid+'" style="display:none;padding:12px;background:#FFF8E1;border-radius:10px;margin-bottom:8px;border:1.5px solid #FFB74D">'
            '<div style="font-size:12px;font-weight:700;color:#E65100;margin-bottom:8px">&#128197; Re-Schedule Meeting</div>'
            '<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:8px">'
            '<input type="date" id="rsd-'+vid+'" class="si">'
            '<input type="time" id="rst-'+vid+'" class="si"></div>'
            '<button onclick="doReschedule(\''+vid+'\',\''+vn+'\',\''+ph+'\')" '
            'style="width:100%;padding:10px;background:#FF6F00;color:white;border:none;border-radius:8px;font-size:13px;font-weight:700;cursor:pointer">'
            '&#128241; Send WhatsApp &amp; Reschedule</button></div>')
    if not pending_cards:
        pending_cards='<div style="text-align:center;padding:18px;color:#999"><div style="font-size:32px">&#10003;</div>No pending requests</div>'

    hist_items=""
    for v in all_visitors:
        ini=(v.get("name","?")[:1]).upper()
        bg="#2E7D32" if v["status"]=="approved" else "#C62828" if v["status"]=="rejected" else "#1565C0" if v["status"]=="rescheduled" else "#F57F17"
        bdg={"approved":"<span class='b-ok'>APPROVED</span>","rejected":"<span class='b-no'>REJECTED</span>","rescheduled":"<span class='b-rs'>RESCHEDULED</span>"}.get(v["status"],"<span class='b-pd'>PENDING</span>")
        hist_items+=('<div class="hr"><div class="hav" style="background:'+bg+'">'+ini+'</div>'
            '<div class="hi"><div class="hn">'+v["name"]+'</div><div class="hs">'+v["purpose"][:30]+'</div></div>'
            '<div class="hrt"><div class="ht">'+v["created_at"][:16]+'</div>'+bdg+'</div></div>')
    if not hist_items:
        hist_items='<div style="text-align:center;padding:18px;color:#999">No history</div>'

    return ("""<!DOCTYPE html><html><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1,user-scalable=no">
<title>"""+name+""" · Maxwell</title>
<style>*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif;background:#F2F4F7;min-height:100vh;padding-bottom:90px;color:#1A1A2E}
.header{background:linear-gradient(135deg,#1565C0,#0D47A1);padding:12px 16px;display:flex;align-items:center;gap:12px;position:sticky;top:0;z-index:200;box-shadow:0 2px 20px rgba(13,71,161,0.4)}
.hdr-logo{height:36px;filter:brightness(0) invert(1)}
.hdr-greet{flex:1}.hdr-hello{font-size:11px;color:rgba(255,255,255,0.75);font-weight:500}
.hdr-name{font-size:16px;font-weight:800;color:white}
.hdr-acts{display:flex;align-items:center;gap:10px}
.nb-btn{width:38px;height:38px;background:rgba(255,255,255,0.15);border-radius:50%;display:flex;align-items:center;justify-content:center;cursor:pointer;border:none;color:white;font-size:18px;position:relative}
.nb-dot{position:absolute;top:6px;right:6px;background:#FF5252;color:white;border-radius:50%;min-width:16px;height:16px;padding:0 3px;font-size:9px;font-weight:800;display:flex;align-items:center;justify-content:center;border:1.5px solid #0D47A1}
.prof-btn{width:38px;height:38px;border-radius:50%;overflow:hidden;cursor:pointer;border:2px solid rgba(255,255,255,0.4)}
.prof-btn img{width:100%;height:100%;object-fit:cover}
.ab{margin:12px 14px 0;background:#FFF8E1;border:1.5px solid #FFB74D;border-radius:14px;padding:12px 16px;display:flex;align-items:center;justify-content:space-between}
.ab-left{display:flex;align-items:center;gap:8px;font-size:13px;font-weight:600;color:#E65100}
.ab-link{font-size:12px;font-weight:700;color:#1565C0;text-decoration:none;border:1.5px solid #1565C0;padding:4px 10px;border-radius:8px}
.nb{margin:12px 14px 0;background:#E3F2FD;border-left:4px solid #1565C0;border-radius:14px;padding:13px 16px;font-size:13px;font-weight:700;color:#0D47A1;display:none}
.pg{padding:14px}
.sc{background:white;border-radius:20px;padding:18px;box-shadow:0 2px 15px rgba(0,0,0,0.07);margin-bottom:14px}
.sc-hdr{display:flex;align-items:center;justify-content:space-between;margin-bottom:14px}
.sc-ttl{display:flex;align-items:center;gap:8px;font-size:15px;font-weight:800;color:#1A1A2E}
.sc-ico{width:32px;height:32px;background:#E3F2FD;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:16px}
.avc{background:#F8FAFE;border-radius:14px;padding:14px;border:1.5px solid #E3F2FD;margin-bottom:10px}
.av-top{display:flex;align-items:flex-start;justify-content:space-between;gap:10px;margin-bottom:12px}
.av-name{font-size:16px;font-weight:800;color:#1A1A2E}.av-time{font-size:11px;color:#888;margin-top:3px}
.av-purpose{font-size:12px;color:#555;margin-top:6px;font-weight:500}.av-left{flex:1}
.co-btn{background:#E53935;color:white;border:none;border-radius:10px;padding:9px 14px;font-size:12px;font-weight:700;cursor:pointer;flex-shrink:0}
.ob{background:white;border-radius:12px;padding:12px 14px;border:1.5px solid #EEF2FF}
.ob-title{font-size:13px;font-weight:700;color:#1565C0;margin-bottom:10px}
.drink-row{display:flex;align-items:center;gap:10px}
.drink-sel{flex:1;border:1.5px solid #E8ECF4;border-radius:10px;padding:10px 12px;font-size:14px;background:#F8F9FA;outline:none;font-family:inherit}
.qty-wrap{display:flex;align-items:center;gap:10px;flex-shrink:0}
.qty-btn{width:36px;height:36px;border-radius:50%;border:2px solid #E0E7FF;background:white;color:#1565C0;cursor:pointer;font-weight:700;font-size:18px}
.qty-num{font-size:18px;font-weight:900;color:#1A1A2E;min-width:24px;text-align:center}
.si{width:100%;background:#F8F9FA;border:1.5px solid #E8ECF4;border-radius:10px;padding:11px 14px;font-size:14px;color:#333;outline:none;font-family:inherit}
.co-btn2{width:100%;margin-top:12px;padding:14px;background:linear-gradient(135deg,#1565C0,#1976D2);color:white;border:none;border-radius:12px;font-size:15px;font-weight:700;cursor:pointer}
.tl{font-size:12px;font-weight:600;color:#F57F17;margin-top:8px;text-align:center}
.sw{background:white;border-radius:20px;padding:14px;box-shadow:0 2px 15px rgba(0,0,0,0.07);margin-bottom:14px}
.sr{display:grid;grid-template-columns:repeat(4,1fr);gap:8px}
.si2{text-align:center;padding:8px 4px}.si2-ico{font-size:22px;margin-bottom:4px}
.si2-num{font-size:20px;font-weight:900;line-height:1}.si2-lbl{font-size:10px;color:#888;font-weight:600;margin-top:3px}
.c1 .si2-num{color:#7C3AED}.c2 .si2-num{color:#2E7D32}.c3 .si2-num{color:#E65100}.c4 .si2-num{color:#1565C0}
.pi{display:flex;align-items:center;gap:12px;padding:12px 0;border-bottom:1px solid #F5F5F5;flex-wrap:wrap}
.pi:last-child{border-bottom:none}.pi-photo{width:44px;height:44px;border-radius:50%;object-fit:cover;border:2px solid #E0E0E0;flex-shrink:0}
.pi-info{flex:1;min-width:0}.pi-name{font-size:14px;font-weight:700;color:#1A1A2E}
.pi-sub{font-size:11px;color:#888;margin-top:2px}.pi-t{font-size:10px;color:#1565C0;margin-top:2px;font-weight:600}
.pi-acts{display:flex;gap:6px;flex-shrink:0}
.pa-ok{width:34px;height:34px;background:#E8F5E9;color:#2E7D32;border:none;border-radius:50%;font-size:16px;font-weight:700;cursor:pointer}
.pa-no{width:34px;height:34px;background:#FFEBEE;color:#C62828;border:none;border-radius:50%;font-size:16px;font-weight:700;cursor:pointer}
.pa-rs{width:34px;height:34px;background:#FFF8E1;color:#E65100;border:none;border-radius:50%;font-size:16px;cursor:pointer}
.hr{display:flex;align-items:center;gap:12px;padding:12px 0;border-bottom:1px solid #F5F5F5}
.hr:last-child{border-bottom:none}.hav{width:40px;height:40px;border-radius:50%;color:white;display:flex;align-items:center;justify-content:center;font-weight:800;font-size:15px;flex-shrink:0}
.hi{flex:1;min-width:0}.hn{font-size:13px;font-weight:700}.hs{font-size:11px;color:#888;margin-top:2px}
.hrt{display:flex;flex-direction:column;align-items:flex-end;gap:4px;flex-shrink:0}.ht{font-size:10px;color:#AAA}
.b-ok{background:#E8F5E9;color:#2E7D32;padding:3px 9px;border-radius:20px;font-size:10px;font-weight:700}
.b-no{background:#FFEBEE;color:#C62828;padding:3px 9px;border-radius:20px;font-size:10px;font-weight:700}
.b-pd{background:#FFF8E1;color:#F57F17;padding:3px 9px;border-radius:20px;font-size:10px;font-weight:700}
.b-rs{background:#E3F2FD;color:#1565C0;padding:3px 9px;border-radius:20px;font-size:10px;font-weight:700}
.ps-form .fgrp{margin-bottom:12px}.ps-form .fgrp label{display:block;font-size:12px;font-weight:600;color:#555;margin-bottom:5px}
.ps-form .fgrp input{width:100%;padding:11px;border:1.5px solid #E8ECF4;border-radius:10px;font-size:14px;outline:none;background:#F8F9FA;font-family:inherit}
.ps-form .fgrp input:focus{border-color:#1565C0;background:white}
.fab{position:fixed;bottom:90px;right:18px;width:54px;height:54px;background:linear-gradient(135deg,#1565C0,#1976D2);border-radius:50%;display:flex;align-items:center;justify-content:center;box-shadow:0 6px 20px rgba(21,101,192,0.45);cursor:pointer;z-index:150;border:none;color:white;text-decoration:none;font-size:24px}
.bnav{position:fixed;bottom:0;left:0;right:0;background:white;display:flex;justify-content:space-around;align-items:center;padding:8px 0 max(16px,env(safe-area-inset-bottom));box-shadow:0 -2px 20px rgba(0,0,0,0.1);z-index:200}
.ni{display:flex;flex-direction:column;align-items:center;gap:3px;text-decoration:none;color:#9E9E9E;font-size:10px;font-weight:600;padding:4px 14px;border-radius:12px;background:none;border:none;cursor:pointer;font-family:inherit}
.ni.active{color:#1565C0}.ni-ico{font-size:22px}
.mo{position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:500;display:none;align-items:flex-end;justify-content:center}
.mo.open{display:flex}.md{background:white;border-radius:24px 24px 0 0;padding:24px;width:100%;max-width:500px;max-height:90vh;overflow-y:auto}
.md-ttl{font-size:18px;font-weight:800;color:#1A1A2E;margin-bottom:20px;text-align:center}
.pp-wrap{text-align:center;margin-bottom:20px;position:relative;display:inline-block;left:50%;transform:translateX(-50%)}
.pp-img{width:90px;height:90px;border-radius:50%;object-fit:cover;border:3px solid #1565C0}
.pp-edit{position:absolute;bottom:0;right:0;background:#1565C0;color:white;border:none;border-radius:50%;width:28px;height:28px;font-size:14px;cursor:pointer}
.sv-btn{width:100%;padding:14px;background:linear-gradient(135deg,#1565C0,#1976D2);color:white;border:none;border-radius:12px;font-size:15px;font-weight:700;cursor:pointer;margin-top:8px}
.cl-btn{width:100%;padding:12px;background:#f5f5f5;color:#555;border:none;border-radius:12px;font-size:14px;font-weight:600;cursor:pointer;margin-top:8px}
.md input{width:100%;padding:12px;border:1.5px solid #ddd;border-radius:10px;font-size:14px;outline:none;font-family:inherit;margin-top:6px;margin-bottom:8px}
.md label{display:block;font-size:13px;font-weight:600;color:#555;margin-top:8px}
</style></head><body>
<div class="header">
  <img src="MLOGO" class="hdr-logo" alt="">
  <div class="hdr-greet">
    <div class="hdr-hello" id="gh">Good Morning,</div>
    <div class="hdr-name">"""+name+"""</div>
  </div>
  <div class="hdr-acts">
    <button class="nb-btn" id="nb-bell">&#128276;<div class="nb-dot" id="nb-cnt" style="display:none">0</div></button>
    <div class="prof-btn" onclick="openMo()"><img src="EPHOTO" id="hdr-pi" alt=""></div>
  </div>
</div>
"""+SOUND_WIDGET+
("""<div class="ab"><div class="ab-left"><span>&#128274;</span><span>Please change your default password!</span></div><a href="/change-password" class="ab-link">Change Now</a></div>""" if force_change else "")+
"""<div class="nb" id="nb-bar"></div>
<div class="pg">
<div class="sc"><div class="sc-hdr"><div class="sc-ttl"><div class="sc-ico">&#128100;</div>Active Visitors</div></div>"""+active_cards+"""</div>
<div class="sw"><div class="sr">
  <div class="si2 c1"><div class="si2-ico">&#128101;</div><div class="si2-num">"""+str(today_count)+"""</div><div class="si2-lbl">Today</div></div>
  <div class="si2 c2"><div class="si2-ico">&#10003;</div><div class="si2-num">"""+str(checked_out)+"""</div><div class="si2-lbl">Out</div></div>
  <div class="si2 c3"><div class="si2-ico">&#128337;</div><div class="si2-num">"""+str(in_house)+"""</div><div class="si2-lbl">In House</div></div>
  <div class="si2 c4"><div class="si2-ico">&#9749;</div><div class="si2-num">"""+str(orders_today)+"""</div><div class="si2-lbl">Orders</div></div>
</div></div>
<div class="sc" id="pend-sc" """+('style="display:none"' if pc==0 else "")+""">
  <div class="sc-hdr"><div class="sc-ttl"><div class="sc-ico">&#9203;</div>Pending Approval</div>
  <span style="background:#FFF3E0;color:#E65100;padding:3px 10px;border-radius:20px;font-size:12px;font-weight:700">"""+str(pc)+"""</span></div>
  """+pending_cards+"""
</div>
<!-- Pre-Schedule Meeting Section -->
<div class="sc">
  <div class="sc-hdr"><div class="sc-ttl"><div class="sc-ico" style="background:#E8F5E9">&#128197;</div>Pre-Schedule Meeting</div></div>
  <div class="ps-form">
    <div class="fgrp"><label>Visitor Name</label><input type="text" id="ps-name" placeholder="Enter visitor name" class="si"></div>
    <div class="fgrp"><label>Mobile Number</label><input type="tel" id="ps-mobile" placeholder="10-digit mobile" class="si"></div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
      <div class="fgrp"><label>Date</label><input type="date" id="ps-date" class="si"></div>
      <div class="fgrp"><label>Time</label><input type="time" id="ps-time" class="si"></div>
    </div>
    <button class="co-btn2" style="margin-top:4px" onclick="scheduleMeeting()">&#128241; Send WhatsApp Invite</button>
    <div id="ps-msg" style="display:none;margin-top:8px;background:#E8F5E9;color:#2E7D32;padding:10px;border-radius:10px;font-weight:600;text-align:center"></div>
  </div>
</div>
<div class="sc"><div class="sc-hdr"><div class="sc-ttl"><div class="sc-ico" style="background:#FFF8E1">&#128203;</div>Recent History</div></div>"""+hist_items+"""</div>
</div>
<!-- Profile Modal -->
<div class="mo" id="mo"><div class="md">
  <div class="md-ttl">&#128100; My Profile</div>
  <div class="pp-wrap">
    <img src="EPHOTO2" id="mo-pi" class="pp-img" alt="">
    <button class="pp-edit" onclick="document.getElementById('pi-file').click()">&#9998;</button>
    <input type="file" id="pi-file" accept="image/*" style="display:none" onchange="handlePP(this)">
  </div>
  <label>Full Name</label><input type="text" id="p-name" value="EMPNAME">
  <label>Email</label><input type="email" id="p-email" value="EMPEMAIL" readonly style="background:#f5f5f5;color:#888">
  <label>Department</label><input type="text" id="p-dept" value="EMPDEPT">
  <label>Designation</label><input type="text" id="p-desig" value="EMPDESIG" placeholder="e.g. HR Executive">
  <button class="sv-btn" onclick="saveProfile()">&#10003; Save Profile</button>
  <button class="cl-btn" onclick="closeMo()">Cancel</button>
  <hr style="margin:16px 0;border:none;border-top:1px solid #eee">
  <button onclick="location.href='/change-password'" style="width:100%;padding:12px;background:#FFF8E1;color:#E65100;border:1.5px solid #FFB74D;border-radius:12px;font-size:14px;font-weight:700;cursor:pointer">&#128274; Change Password</button>
  <button onclick="location.href='/employee-logout'" style="width:100%;padding:12px;background:#FFEBEE;color:#C62828;border:1.5px solid #EF9A9A;border-radius:12px;font-size:14px;font-weight:700;cursor:pointer;margin-top:8px">&#128682; Logout</button>
</div></div>
<a href="/" class="fab">+</a>
<nav class="bnav">
  <button class="ni active"><div class="ni-ico">&#127968;</div><span>Home</span></button>
  <button class="ni" onclick="showPend()"><div class="ni-ico">&#128101;</div><span>Visitors</span></button>
  <button class="ni" onclick="location.href='/pantry'"><div class="ni-ico">&#9749;</div><span>Pantry</span></button>
  <button class="ni" onclick="location.href='/'"><div class="ni-ico">&#128203;</div><span>Form</span></button>
  <button class="ni" onclick="openMo()"><div class="ni-ico">&#128100;</div><span>Profile</span></button>
</nav>
<script>"""+BEEP_JS+"""
var _qty={},_lv="""+str(lhp_id)+""",_hwr=true,_nc=0,_rm={},_npd=null;
var _vt="""+json.dumps(visitor_times)+""";
var h=new Date().getHours();document.getElementById('gh').textContent=h<12?'Good Morning,':h<17?'Good Afternoon,':'Good Evening,';
function changeQty(v,d){if(!_qty[v])_qty[v]=1;_qty[v]=Math.max(1,_qty[v]+d);document.getElementById('qty-'+v).textContent=_qty[v];}
function parseIST(s){if(!s)return new Date();var p=s.split(' ');var dp=p[0].split('-');var tp=(p[1]||'0:0').split(':');return new Date(parseInt(dp[2]),parseInt(dp[1])-1,parseInt(dp[0]),parseInt(tp[0]),parseInt(tp[1]));}
function chkReveal(){var now=new Date();Object.keys(_vt).forEach(function(vid){
  var ot=parseIST(_vt[vid]);var diff=Math.floor((now-ot)/60000);
  var os=document.getElementById('ord-'+vid);var tl=document.getElementById('tl-'+vid);
  if(os){if(diff>=7){os.style.display='block';if(tl)tl.textContent='';}else{os.style.display='none';if(tl)tl.textContent='Order in '+(7-diff)+' min';}}
  if(diff>=20&&!_rm[vid]){_rm[vid]=true;_beep(4);showNB('&#9749; Please order tea/coffee for your guest!',15000);}});}
async function confirmOrder(vid,vname,person){
  var drink=document.getElementById('drk-'+vid).value;var qty=_qty[vid]||1;
  var snacks=document.getElementById('snk-'+vid).value;var note=document.getElementById('nte-'+vid).value;
  if(!drink&&!snacks){alert('Please select drink or snacks!');return;}
  await fetch('/api/beverage',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({visitor_id:vid,visitor_name:vname,person_to_meet:person,drink:drink,quantity:qty,snacks:snacks,note:note})});
  showNB('&#10003; Order sent!',5000);_beep(2);document.getElementById('ord-'+vid).style.display='none';}
async function act(id,action){if(!confirm(action+' this visitor?'))return;
  await fetch('/action/'+id+'/'+action,{headers:{'Accept':'application/json'}});
  if(action==='approve')window.open('/pass/'+id);location.reload();}
async function checkout(id){if(!confirm('Checkout?'))return;await fetch('/api/checkout/'+id,{method:'POST'});location.reload();}
function toggleRs(vid){var p=document.getElementById('rs-'+vid);p.style.display=p.style.display==='none'?'block':'none';}
async function doReschedule(vid,vname,phone){
  var date=document.getElementById('rsd-'+vid).value;var time=document.getElementById('rst-'+vid).value;
  if(!date||!time){alert('Please select date and time!');return;}
  var dateObj=new Date(date);var dd=dateObj.toLocaleDateString('en-IN',{day:'2-digit',month:'long',year:'numeric'});
  var msg='%F0%9F%94%84 Meeting Rescheduled%0A%0ADear '+encodeURIComponent(vname)+',%0A%0AYour meeting has been rescheduled.%0A%0ANew Date: '+encodeURIComponent(dd)+'%0ANew Time: '+encodeURIComponent(time)+'%0AHost: """+name+"""%0A%0APlease confirm.%0A%0AMaxwell Engineering Solutions';
  await fetch('/api/reschedule/'+vid,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({date:date,time:time})});
  window.open('https://wa.me/91'+phone.replace(/\\D/g,'')+' ?text='+msg);
  showNB('&#128197; Rescheduled & WhatsApp sent!',6000);setTimeout(function(){location.reload();},2000);}
async function scheduleMeeting(){
  var n=document.getElementById('ps-name').value.trim();var m=document.getElementById('ps-mobile').value.trim();
  var d=document.getElementById('ps-date').value;var t=document.getElementById('ps-time').value;
  if(!n||!m||!d||!t){alert('Please fill all fields!');return;}
  var dateObj=new Date(d);var dd=dateObj.toLocaleDateString('en-IN',{day:'2-digit',month:'long',year:'numeric'});
  var msg='%F0%9F%93%85 Meeting Scheduled%0A%0ADear '+encodeURIComponent(n)+',%0A%0AYou have a scheduled meeting at Maxwell Engineering Solutions.%0A%0ADate: '+encodeURIComponent(dd)+'%0ATime: '+encodeURIComponent(t)+'%0AHost: """+name+"""%0A%0APlease arrive 10 minutes before.%0A%0AMaxwell Engineering Solutions';
  await fetch('/api/schedule-meeting',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({host_name:'"""+name+"""',visitor_name:n,visitor_phone:m,meeting_date:d,meeting_time:t})});
  window.open('https://wa.me/91'+m.replace(/\\D/g,'')+' ?text='+msg);
  var me=document.getElementById('ps-msg');me.textContent='\\u2713 WhatsApp opened! Invite sent to '+n;me.style.display='block';
  setTimeout(function(){me.style.display='none';},5000);
  document.getElementById('ps-name').value='';document.getElementById('ps-mobile').value='';
  document.getElementById('ps-date').value='';document.getElementById('ps-time').value='';}
function showNB(msg,dur){var b=document.getElementById('nb-bar');b.innerHTML=msg;b.style.display='block';setTimeout(function(){b.style.display='none';},dur||8000);}
function addNC(){_nc++;var e=document.getElementById('nb-cnt');e.style.display='flex';e.textContent=_nc;}
function showPend(){document.getElementById('pend-sc').style.display='block';document.getElementById('pend-sc').scrollIntoView({behavior:'smooth'});}
function openMo(){document.getElementById('mo').classList.add('open');}
function closeMo(){document.getElementById('mo').classList.remove('open');_npd=null;}
function handlePP(input){if(!input.files||!input.files[0])return;var r=new FileReader();r.onload=function(e){_npd=e.target.result;document.getElementById('mo-pi').src=_npd;document.getElementById('hdr-pi').src=_npd;};r.readAsDataURL(input.files[0]);}
async function saveProfile(){var pld={name:document.getElementById('p-name').value.trim(),department:document.getElementById('p-dept').value.trim(),designation:document.getElementById('p-desig').value.trim()};
  if(_npd)pld.photo=_npd;var r=await fetch('/api/profile',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(pld)});
  var d=await r.json();if(d.success){showNB('&#10003; Profile saved!',4000);closeMo();}else{alert('Error!');}}
async function checkNew(){try{
  var r0=await fetch('/api/latest-pending?host='+encodeURIComponent('"""+name+"""'));var d0=await r0.json();
  if(_hwr&&d0.visitor&&d0.visitor.id>_lv){_beep(4);addNC();showNB('&#128276; New visitor: '+d0.visitor.name,15000);}
  if(d0.visitor)_lv=d0.visitor.id;_hwr=true;
  var r1=await fetch('/api/pending-count');var d1=await r1.json();
  document.title=d1.count>0?'('+d1.count+') \\u00B7 """+name+"""':'"""+name+""" \\u00B7 Maxwell';}catch(e){}}
if(Notification.permission==='default')Notification.requestPermission();
setInterval(checkNew,8000);checkNew();setInterval(chkReveal,15000);chkReveal();
</script></body></html>""").replace("MLOGO",LOGO_MAIN).replace("EPHOTO",emp_photo).replace("EPHOTO2",emp_photo).replace("EMPNAME",name).replace("EMPEMAIL",email).replace("EMPDEPT",emp_dept).replace("EMPDESIG",emp_desig)

@app.route("/employee-logout")
def employee_logout():
    session.clear(); return redirect("/employee-login")

@app.route("/change-password", methods=["GET","POST"])
def change_password():
    if not session.get("emp_name"): return redirect("/employee-login")
    msg=""; err=""
    if request.method=="POST":
        old_pw=request.form.get("old_password",""); new_pw=request.form.get("new_password",""); confirm_pw=request.form.get("confirm_password","")
        email=session.get("emp_email",""); emp_name=session.get("emp_name","")
        pw_ok,_=check_emp_password(email,old_pw,emp_name)
        if not pw_ok: err="Current password is wrong!"
        elif len(new_pw)<4: err="New password must be at least 4 characters!"
        elif new_pw!=confirm_pw: err="Passwords do not match!"
        else:
            conn=get_db();conn.execute("INSERT OR REPLACE INTO employee_passwords (email,password_hash,is_default) VALUES (?,?,0)",(email,hash_pw(new_pw)));conn.commit();conn.close()
            session.pop("force_pw_change",None); msg="Password changed successfully!"
    return ("<!DOCTYPE html><html><head><meta charset='UTF-8'><title>Change Password</title>"
            "<style>body{font-family:Arial;background:#f0f4f8}"+HEADER_CSS
            +".box{max-width:420px;margin:40px auto;background:white;padding:32px;border-radius:13px;box-shadow:0 5px 18px rgba(0,0,0,0.1)}"
            ".box h2{color:#1565C0;margin-bottom:22px;text-align:center}label{display:block;font-size:13px;font-weight:600;color:#444;margin-bottom:5px;margin-top:14px}"
            "input{width:100%;padding:11px;border:2px solid #e0e0e0;border-radius:7px;font-size:14px}input:focus{outline:none;border-color:#1565C0}"
            "button{width:100%;margin-top:18px;padding:12px;background:#1565C0;color:white;border:none;border-radius:7px;font-size:15px;font-weight:700;cursor:pointer}"
            ".err{color:red;font-size:13px;margin-top:8px;background:#ffebee;padding:10px;border-radius:7px;text-align:center}"
            ".msg{color:#2E7D32;font-size:13px;margin-top:8px;background:#E8F5E9;padding:10px;border-radius:7px;font-weight:600;text-align:center}"
            "</style></head><body>"
            +make_header(LOGO_MAIN,'<a href="/employee-dashboard">&#8592; Back</a>')+
            "<div class='box'><h2>&#128274; Change Password</h2><form method='POST'>"
            "<label>Current Password</label><input type='password' name='old_password' required>"
            "<label>New Password</label><input type='password' name='new_password' required>"
            "<label>Confirm New Password</label><input type='password' name='confirm_password' required>"
            "<button type='submit'>Update Password</button></form>"
            +(("<p class='err'>"+err+"</p>") if err else "")
            +(("<p class='msg'>&#10003; "+msg+" <a href='/employee-dashboard'>Back</a></p>") if msg else "")
            +"</div></body></html>")

# ══════════════════════════════════════════════════
# SECURITY (with Exit feature)
# ══════════════════════════════════════════════════
@app.route("/security-login", methods=["GET","POST"])
def security_login():
    err=""
    if request.method=="POST":
        mobile=request.form.get("mobile","").strip(); password=request.form.get("password","").strip()
        if mobile==SECURITY_MOBILE and password==get_setting("security_pass","1234"):
            session["security_ok"]=True; return redirect("/security-dashboard")
        err="Wrong mobile or password!"
    return ("<!DOCTYPE html><html><head><meta charset='UTF-8'><title>Security Login</title>"
            "<style>body{font-family:Arial;background:#f0f4f8}.box{max-width:400px;margin:80px auto;background:white;padding:38px;border-radius:13px;box-shadow:0 5px 18px rgba(0,0,0,0.1)}"
            ".box h2{color:#1565C0;text-align:center;margin-bottom:22px}label{display:block;font-size:13px;font-weight:600;color:#444;margin-bottom:5px;margin-top:14px}"
            "input{width:100%;padding:11px;border:2px solid #e0e0e0;border-radius:7px;font-size:14px}input:focus{outline:none;border-color:#1565C0}"
            "button{width:100%;margin-top:18px;padding:12px;background:#1565C0;color:white;border:none;border-radius:7px;font-size:15px;font-weight:700;cursor:pointer}"
            ".err{color:red;font-size:13px;margin-top:8px;text-align:center}"
            "</style></head><body><div class='box'><h2>&#128110; Security Login</h2>"
            "<form method='POST'><label>Mobile Number</label><input type='tel' name='mobile' placeholder='9023730509' required>"
            "<label>Password</label><input type='password' name='password' required>"
            "<button type='submit'>LOGIN</button></form>"
            +(("<p class='err'>"+err+"</p>") if err else "")
            +"<p style='text-align:center;margin-top:12px'><a href='/' style='color:#1565C0'>Back to Form</a></p>"
            "</div></body></html>")

@app.route("/security-dashboard")
def security_dashboard():
    if not session.get("security_ok"): return redirect("/security-login")
    conn=get_db()
    visitors=[dict(r) for r in conn.execute(
        "SELECT id,name,phone,category,department,person_to_meet,status,created_at,checkout_at,exit_at,pass_number,photo FROM visitors ORDER BY id DESC LIMIT 80"
    ).fetchall()]
    latest_id=visitors[0]["id"] if visitors else 0
    lco=conn.execute("SELECT id FROM visitors WHERE checkout_at IS NOT NULL ORDER BY id DESC LIMIT 1").fetchone()
    lco_id=lco["id"] if lco else 0
    conn.close()
    rows=""
    for v in visitors:
        bc=v["status"] if v["status"] in ("pending","approved","rejected","rescheduled") else "pending"
        photo=v["photo"] if v.get("photo") else DEFAULT_PHOTO
        pass_btn=""; exit_btn=""
        if v["status"]=="approved":
            pass_btn='<button class="btn bp" onclick="window.open(\'/pass/'+str(v["id"])+'\')">&#128203; Pass</button>'
        # Show EXIT button only when checked out but not yet exited
        if v.get("checkout_at") and not v.get("exit_at"):
            exit_btn='<button class="btn" style="background:#4A148C;color:white" onclick="secExit('+str(v["id"])+')">&#128682; Exit</button>'
        co=v.get("checkout_at") or "-"; ex=v.get("exit_at") or "-"
        ex_color="#2E7D32" if v.get("exit_at") else ("#C62828" if v.get("checkout_at") else "#888")
        rows+=("<tr><td><img src='"+photo+"' style='width:38px;height:38px;border-radius:50%;object-fit:cover;border:2px solid #ddd'></td>"
               "<td><b>"+str(v["name"])+"</b></td><td>"+str(v["phone"])+"</td>"
               "<td>"+str(v["department"])+"</td><td>"+str(v["person_to_meet"])+"</td>"
               "<td style='font-size:11px'>"+str(v["created_at"])+"</td>"
               "<td style='font-size:11px'>"+co+"</td>"
               "<td style='font-size:11px;color:"+ex_color+"'>"+ex+"</td>"
               "<td><span class='badge "+bc+"'>"+v["status"].upper()+"</span></td>"
               "<td>"+pass_btn+exit_btn+"</td></tr>")
    if not rows: rows='<tr><td colspan="10" style="text-align:center;padding:25px;color:#999">No visitors</td></tr>'
    hdr=make_header(LOGO_MAIN,'<a href="/">Visitor Form</a><a href="/security-logout">Logout</a>')
    return ("<!DOCTYPE html><html><head><meta charset='UTF-8'><title>Security Dashboard</title>"
            "<style>*{box-sizing:border-box;margin:0;padding:0}body{font-family:Arial;background:#f0f4f8}"+HEADER_CSS
            +".container{max-width:1150px;margin:20px auto;padding:0 15px}"
            ".card{background:white;border-radius:9px;padding:18px;box-shadow:0 2px 8px rgba(0,0,0,0.07);overflow-x:auto;margin-bottom:15px}"
            ".card h3{color:#1565C0;margin-bottom:14px}"
            ".nb{background:#FFF8E1;border:2px solid #F57F17;border-radius:9px;padding:12px 18px;margin-bottom:15px;display:none;font-weight:700;color:#E65100;font-size:14px}"
            "table{width:100%;border-collapse:collapse;min-width:850px}"
            "th{background:#1565C0;color:white;padding:10px;font-size:12px;text-align:left}"
            "td{padding:9px;border-bottom:1px solid #eee;font-size:13px;vertical-align:middle}tr:hover td{background:#f9f9f9}"
            ".badge{display:inline-block;padding:3px 9px;border-radius:11px;font-size:11px;font-weight:700}"
            ".badge.pending{background:#FFF8E1;color:#F57F17}.badge.approved{background:#E8F5E9;color:#2E7D32}"
            ".badge.rejected{background:#FFEBEE;color:#C62828}.badge.rescheduled{background:#E3F2FD;color:#1565C0}"
            ".btn{padding:5px 10px;border:none;border-radius:5px;cursor:pointer;font-size:11px;font-weight:600;margin:2px}.bp{background:#1565C0;color:white}"
            ".legend{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:12px;font-size:12px;color:#555}"
            ".ld{display:flex;align-items:center;gap:5px}.ldot{width:10px;height:10px;border-radius:50%}"
            "</style></head><body>"+hdr+SOUND_WIDGET+
            "<div class='container'><div class='nb' id='nb'></div>"
            "<div class='card'><h3>&#128100; Visitor Entries</h3>"
            "<div class='legend'>"
            "<div class='ld'><div class='ldot' style='background:#2E7D32'></div>Exit Time = Confirmed Exit</div>"
            "<div class='ld'><div class='ldot' style='background:#C62828'></div>Red Exit = Checked out, needs to exit</div>"
            "<div class='ld'><div class='ldot' style='background:#888'></div>Grey = Not checked out yet</div>"
            "</div>"
            "<table><tr><th>Photo</th><th>Name</th><th>Phone</th><th>Dept</th><th>Meeting</th><th>In Time</th><th>Checkout</th><th>Exit Time</th><th>Status</th><th>Action</th></tr>"
            +rows+"</table></div></div>"
            "<script>"+BEEP_JS
            +"var _lv="+str(latest_id)+",_lco="+str(lco_id)+";"
            "function showNB(msg,dur){var b=document.getElementById('nb');b.innerHTML=msg;b.style.display='block';setTimeout(function(){b.style.display='none';},dur||8000);}"
            "async function secExit(id){"
            "if(!confirm('Confirm visitor has physically exited the premises?'))return;"
            "var r=await fetch('/api/security-exit/'+id,{method:'POST'});"
            "var d=await r.json();"
            "if(d.success){showNB('&#10003; Exit confirmed at '+d.exit_at,8000);setTimeout(function(){location.reload();},2000);}}"
            "async function checkNew(){try{"
            "var r=await fetch('/api/latest-visitor');var d=await r.json();"
            "if(d.visitors&&d.visitors.length>0&&d.visitors[0].id>_lv){"
            "_beep(4);showNB('&#128276; New visitor: '+d.visitors[0].name,10000);_lv=d.visitors[0].id;setTimeout(function(){location.reload();},3000);}"
            "var r2=await fetch('/api/checkout-notify');var d2=await r2.json();"
            "if(d2.checkout&&d2.checkout.id>_lco){"
            "_beep(3);showNB('&#128682; '+d2.checkout.name+' checked out \\u2014 please allow exit!',12000);_lco=d2.checkout.id;"
            "setTimeout(function(){location.reload();},3000);}}"
            "catch(e){}}"
            "setInterval(checkNew,8000);checkNew();"
            "</script></body></html>")

@app.route("/security-logout")
def security_logout():
    session.pop("security_ok",None); return redirect("/security-login")

@app.route("/manifest.json")
def manifest():
    return jsonify({"name":"Maxwell Visitor Management","short_name":"Maxwell VM","start_url":"/","display":"standalone","background_color":"#1565C0","theme_color":"#1565C0","icons":[{"src":"/icon.png","sizes":"192x192","type":"image/png","purpose":"any maskable"}]})

@app.route("/sw.js")
def service_worker():
    from flask import Response as R
    sw=("const CACHE='maxwell-v3';self.addEventListener('install',e=>{self.skipWaiting();});"
        "self.addEventListener('activate',e=>{self.clients.claim();});"
        "self.addEventListener('fetch',e=>{if(e.request.method!=='GET')return;"
        "e.respondWith(fetch(e.request).catch(()=>caches.match(e.request)));});")
    return R(sw,mimetype="application/javascript")

@app.route("/icon.png")
def app_icon():
    from flask import Response as R
    svg=b'<svg xmlns="http://www.w3.org/2000/svg" width="192" height="192"><rect width="192" height="192" fill="#1565C0" rx="20"/><text x="96" y="130" font-family="Arial" font-size="100" font-weight="bold" fill="white" text-anchor="middle">M</text></svg>'
    return R(svg,mimetype="image/svg+xml")

init_db()

if __name__ == "__main__":
    print("\n" + "="*55)
    print("  Maxwell Visitor Management System v3.0")
    print("="*55)
    print("  App:      http://localhost:5000")
    print("  Admin:    http://localhost:5000/admin  PIN: 1234")
    print("  Pantry:   http://localhost:5000/pantry")
    print("  Security: http://localhost:5000/security-login")
    print("  Employee: http://localhost:5000/employee-login")
    print("="*55 + "\n")
    app.run(debug=True, host="0.0.0.0", port=5000)
